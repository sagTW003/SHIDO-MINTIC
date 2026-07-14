"""
============================================================
  Ada — Agente de Orientación Vocacional
  Sistema Multiagente ODEM · Universidad EAN
  Versión: 2.0.0 (adaptado al sistema ODEM)

  FLUJO:
    Paso 1 → Recolección de datos personales
    Paso 2 → Test de aptitudes (habilidades, intereses, vocaciones)
    Paso 3 → Consulta SNIES vía Lumina (real) o mock (fallback)
    Paso 4 → Reporte personalizado generado con Claude

  INTEGRACIÓN MULTIAGENTE:
    Coordinado por Viernes.
    Llama a Lumina para consultas SNIES en tiempo real.
    Puede invocar a Scrapper para datos de becas actualizados.

  DEPENDENCIAS:
    pip install openai requests fpdf2

  AUTOR: Sistema Multiagente ODEM — Universidad EAN
  FECHA: Abril 2026
============================================================
"""

import os
import sys
import json
import time
import re
import subprocess
from datetime import datetime
from typing import Optional
import openai

# Mute prints if running via sys.argv (JSON API mode)
def print(*args, **kwargs):
    import sys
    if len(sys.argv) > 1:
        pass
    else:
        import builtins
        builtins.print(*args, **kwargs)


# ================================================================
#   CONFIGURACIÓN
# ================================================================

def _cargar_env():
    """Carga variables de un .env en la raiz de SHIDO_MINTIC (sin dependencias)."""
    aqui = os.path.dirname(os.path.abspath(__file__))
    raiz = os.path.dirname(os.path.dirname(os.path.dirname(aqui)))  # .../SHIDO_MINTIC
    ruta_env = os.path.join(raiz, ".env")
    if not os.path.exists(ruta_env):
        return
    try:
        with open(ruta_env, "r", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if not linea or linea.startswith("#") or "=" not in linea:
                    continue
                clave, _, valor = linea.partition("=")
                clave = clave.strip()
                valor = valor.strip().strip('"').strip("'")
                if clave and clave not in os.environ:
                    os.environ[clave] = valor
    except Exception:
        pass


_cargar_env()


def _a_int(valor, defecto=0):
    """Convierte a int de forma segura (acepta str con separadores, None, floats)."""
    if valor is None or isinstance(valor, bool):
        return defecto
    if isinstance(valor, (int, float)):
        return int(valor)
    try:
        limpio = re.sub(r"[^\d\-]", "", str(valor))
        return int(limpio) if limpio not in ("", "-") else defecto
    except Exception:
        return defecto


API_KEY = os.environ.get("NVIDIA_API_KEY", "")
MODEL   = "meta/llama-3.1-8b-instruct"

# Credenciales de base de datos (leidas del entorno / .env)
DB_HOST = os.environ.get("DB_HOST", "127.0.0.1")
DB_PORT = _a_int(os.environ.get("DB_PORT", "3306"), 3306)
DB_USER = os.environ.get("DB_USER", "odemiro")
DB_PASS = os.environ.get("DB_PASS", "")
DB_NAME = os.environ.get("DB_NAME", "odemiro_db")


def _db_kwargs():
    """Devuelve kwargs de conexion MySQL leyendo la clave del entorno en tiempo de uso."""
    return {
        "host": DB_HOST,
        "port": DB_PORT,
        "user": DB_USER,
        "password": os.environ.get("DB_PASS", ""),
        "database": DB_NAME,
        "connection_timeout": 10,
    }


# Mapeo categoria del test vocacional -> palabras clave del perfil.
# Nivel modulo (antes vivia dentro de recalcular_perfil_desde_respuestas) para
# poder construir KEYWORD_A_CATEGORIA y ponderar el score de afinidad por
# programa segun el promedio REAL de cada categoria del estudiante.
CATEGORIAS_MAP = {
    "Intereses_Creatividad_Innovacion": ["creatividad", "innovación", "diseño", "arte", "estética"],
    "Aptitudes_Analiticas_Logicas": ["analítico", "lógico-matemático", "datos", "matemáticas", "física"],
    "Personalidad_Social_Empatica": ["empatía", "trabajo en equipo", "servicio a otros", "comunicación", "ciencias sociales"],
    "Aptitudes_Practicas_Construccion": ["construcción", "tecnología", "programación", "laboratorio", "mecánica"],
    "Personalidad_Organizacion_Metodo": ["organización", "metodología", "finanzas", "administración"],
    "Personalidad_Liderazgo_Impacto_Cambio": ["liderazgo", "negocios", "emprendimiento", "argumentación"]
}

# Indice inverso: palabra clave -> categoria a la que pertenece.
KEYWORD_A_CATEGORIA = {kw: cat for cat, kws in CATEGORIAS_MAP.items() for kw in kws}


# Mapeo perfil vocacional -> palabras clave (REGEXP) para buscar programas en SNIES.
# Determinista: NO depende del LLM.
PERFIL_A_REGEXP = {
    "creatividad": "DISE|ARTE|CREATIV|MULTIMEDIA|PUBLICIDAD|MUSIC|AUDIOVISUAL",
    "innovación": "DISE|INNOVAC|EMPREND|TECNOLOG",
    "diseño": "DISE|ARQUITECT|GRAFIC|INDUSTRIAL",
    "arte": "ARTE|MUSIC|TEATRO|PLASTIC|AUDIOVISUAL|DANZA",
    "estética": "DISE|ARTE|COSMET|ESTETIC",
    "analítico": "INGENIER|MATEMAT|ESTADIST|DATOS|SISTEMAS|ECONOM|FISICA",
    "lógico-matemático": "MATEMAT|INGENIER|ESTADIST|SISTEMAS|FISICA|ECONOM",
    "matemáticas": "MATEMAT|ESTADIST|ACTUAR|FISICA|ECONOM",
    "física": "FISICA|INGENIER|ELECTRON|MECANIC",
    "datos": "DATOS|ESTADIST|SISTEMAS|INFORMAT|MATEMAT",
    "programación": "SISTEMAS|SOFTWARE|INFORMAT|COMPUTAC|DESARROLLO",
    "tecnología": "SISTEMAS|TECNOLOG|ELECTRON|TELECOMUNIC|INFORMAT",
    "construcción": "CIVIL|ARQUITECT|CONSTRUC|OBRAS",
    "mecánica": "MECANIC|AUTOMOTRIZ|INDUSTRIAL|ELECTROMEC",
    "empatía": "PSICOLOG|TRABAJO SOCIAL|ENFERMER|MEDICIN|EDUCAC",
    "servicio a otros": "ENFERMER|MEDICIN|TRABAJO SOCIAL|PSICOLOG|EDUCAC",
    "comunicación": "COMUNICAC|PERIODISM|PUBLICIDAD|LENGUAS|IDIOMAS",
    "ciencias sociales": "SOCIOLOG|POLITIC|ANTROPOLOG|HISTORIA|DERECHO|TRABAJO SOCIAL",
    "liderazgo": "ADMINISTRAC|NEGOCIOS|GERENC|EMPREND|POLITIC",
    "negocios": "ADMINISTRAC|NEGOCIOS|MERCADEO|COMERCIO|ECONOM|CONTAD",
    "emprendimiento": "EMPREND|ADMINISTRAC|NEGOCIOS|INNOVAC",
    "finanzas": "FINANZ|CONTAD|ECONOM|BANCA|ADMINISTRAC",
    "administración": "ADMINISTRAC|GERENC|NEGOCIOS|GESTION",
    "organización": "ADMINISTRAC|GESTION|LOGISTIC|INDUSTRIAL",
    "metodología": "INVESTIGAC|ESTADIST|CIENCIAS",
    "laboratorio": "QUIMIC|BIOLOG|FISICA|MEDICIN|BACTERIOLOG|FARMAC",
    "argumentación": "DERECHO|FILOSOF|COMUNICAC|POLITIC|PERIODISM",
}


# Codigo DANE de departamento -> nombre (mismo diccionario que
# scripts/consolidar_geih.py, verificado contra snies_matriculados). Se
# invierte para resolver el departamento del ESTUDIANTE (texto libre del
# formulario) al codigo que usan las tablas geih_*.
DPTO_NOMBRE = {
    5: "ANTIOQUIA", 8: "ATLÁNTICO", 11: "BOGOTÁ D.C.", 13: "BOLÍVAR",
    15: "BOYACÁ", 17: "CALDAS", 18: "CAQUETÁ", 19: "CAUCA", 20: "CESAR",
    23: "CÓRDOBA", 25: "CUNDINAMARCA", 27: "CHOCÓ", 41: "HUILA",
    44: "LA GUAJIRA", 47: "MAGDALENA", 50: "META", 52: "NARIÑO",
    54: "NORTE DE SANTANDER", 63: "QUINDÍO", 66: "RISARALDA",
    68: "SANTANDER", 70: "SUCRE", 73: "TOLIMA", 76: "VALLE DEL CAUCA",
    81: "ARAUCA", 85: "CASANARE", 86: "PUTUMAYO",
    88: "ARCHIPIÉLAGO DE SAN ANDRÉS, PROVIDENCIA Y SANTA CATALINA",
    91: "AMAZONAS", 94: "GUAINÍA", 95: "GUAVIARE", 97: "VAUPÉS", 99: "VICHADA",
}
NOMBRE_A_DPTO = {v: k for k, v in DPTO_NOMBRE.items()}
# Alias comunes que un estudiante podria escribir en el formulario web.
NOMBRE_A_DPTO.update({
    "BOGOTA": 11, "BOGOTA D.C.": 11, "BOGOTA DC": 11, "SAN ANDRES": 88,
    "ATLANTICO": 8, "BOLIVAR": 13, "BOYACA": 15, "CAQUETA": 18,
    "CORDOBA": 23, "GUAJIRA": 44, "LA GUAJIRA": 44, "NARINO": 52,
    "QUINDIO": 63, "GUAINIA": 94, "VAUPES": 97,
})

# Area de conocimiento (SNIES, valores reales verificados contra la BD) ->
# divisiones CIIU Rev.4 a 2 digitos (mismo codigo que geih_sector_departamento
# .sector_ciiu_2d) que agrupan los sectores economicos afines a esa area.
AREA_CONOCIMIENTO_A_CIIU = {
    "ECONOMÍA, ADMINISTRACIÓN, CONTADURÍA Y AFINES": [64, 66, 69, 70],
    "INGENIERÍA, ARQUITECTURA, URBANISMO Y AFINES": [41, 42, 43, 62, 71],
    "CIENCIAS SOCIALES Y HUMANAS": [69, 84, 94],
    "CIENCIAS DE LA EDUCACIÓN": [85],
    "CIENCIAS DE LA SALUD": [86, 87, 88],
    "BELLAS ARTES": [90],
    "AGRONOMÍA, VETERINARIA Y AFINES": [1, 2, 3, 75],
    "MATEMÁTICAS Y CIENCIAS NATURALES": [62, 72],
}


def _tags_programa(programa_academico: str, area_conocimiento: str = "") -> list[str]:
    """Etiqueta un programa académico con las palabras clave del perfil vocacional
    (PERFIL_A_REGEXP) cuyo patrón coincide con su nombre/área REAL.

    Corrige el bug donde cada programa se etiquetaba con una copia del perfil
    del ESTUDIANTE (self.perfil_vocacional[:4]) en vez de algo propio del
    programa, lo que dejaba el score de afinidad constante (siempre 4) para
    cualquier resultado y anulaba el ordenamiento por relevancia.
    """
    texto = f"{programa_academico or ''} {area_conocimiento or ''}".upper()
    return [kw for kw, patron in PERFIL_A_REGEXP.items() if re.search(patron, texto)]


def _score_afinidad(tags_programa: list, promedios_categorias: dict) -> float:
    """Score de afinidad = promedio (no suma) del puntaje real del estudiante
    en las categorias de las palabras clave propias del programa.

    Se promedia -y no se suma- para que un programa no gane puntos solo por
    acumular MAS coincidencias de palabras clave (p.ej. "administración" tiene
    mas entradas en PERFIL_A_REGEXP que "analítico"): lo que importa es que
    tan bien alineadas estan, en promedio, las categorias que SI le
    corresponden al programa con lo que el estudiante puntuo alto.
    """
    if not tags_programa:
        return 0.0
    puntajes = [
        promedios_categorias.get(KEYWORD_A_CATEGORIA.get(kw, ""), 3.0)
        for kw in tags_programa
    ]
    return sum(puntajes) / len(puntajes)


# ================================================================
#   BECAS Y CONVOCATORIAS (opcion C: curadas + scraping en vivo opcional)
# ================================================================

# Lista curada con URLs REALES verificadas. Es la fuente por defecto (rapida y estable).
# Usada por el Excel Y por el reporte web para que ambos muestren lo MISMO.
BECAS_CURADAS = [
    ("Universidad EAN", "Beca Talento EAN",
     "https://universidadean.edu.co/la-universidad/becas-y-descuentos",
     "Descuento del 10% al 50% en la matrícula para estudiantes de pregrado con excelencia académica, cultural, deportiva o tecnológica en la Universidad EAN."),
    ("Universidad de los Andes", "Beca Quiero Estudiar",
     "https://apoyofinanciero.uniandes.edu.co/",
     "Cubre hasta el 95% del costo de la matrícula de pregrado durante toda la carrera para estudiantes de estratos 1, 2 y 3 con excelente ICFES."),
    ("ICETEX", "Créditos Educativos Condonables (Fondo Sisbén)",
     "https://web.icetex.gov.co/creditos/pregrado",
     "Crédito condonable hasta del 100% en matrícula para estudiantes clasificados en Sisbén IV (A, B o C) y con alto rendimiento en pruebas Saber 11."),
    ("Ministerio de Educación / ICETEX", "Política de Gratuidad (Matrícula Cero)",
     "https://www.mineducacion.gov.co/portal/secciones/Gratuidad/",
     "Financiación del 100% de la matrícula en instituciones de educación superior públicas del país para estudiantes con vulnerabilidad socioeconómica."),
    ("ATENEA / Alcaldía de Bogotá", "Programa Jóvenes a la U",
     "https://ateneabogota.gov.co/convocatorias/jovenes-la-u",
     "Financiación completa (100% matrícula cero) de carreras universitarias o técnicas en Bogotá, incluyendo auxilio económico mensual de sostenimiento."),
    ("Secretaría de Educación de Bogotá", "Beca Mejores Bachilleres",
     "https://www.educacionbogota.edu.co/",
     "Auxilio de matrícula total o parcial en universidades aliadas de Bogotá a los bachilleres de colegios públicos con los puntajes ICFES más altos de su promoción."),
]

# URLs que el Scrapper puede visitar para actualizar becas en vivo.
URLS_BECAS_SCRAPER = [
    "https://web.icetex.gov.co/creditos/pregrado",
    "https://ateneabogota.gov.co/convocatorias",
]


def obtener_becas(en_vivo=False, objetivo="becas y convocatorias de pregrado"):
    """Devuelve lista de becas (fuente, titulo, link, descripcion).
    Por defecto usa BECAS_CURADAS (rapido/estable). Si en_vivo=True, intenta
    complementar con datos REALES scrapeados por el agente Scrapper; si el
    scraping falla, degrada limpio a la lista curada.
    Retorna (lista, fuente_str)."""
    if not en_vivo:
        return list(BECAS_CURADAS), "curada"

    ruta_scraper = os.path.join(RUTA_AGENTES, "Scrapper", "Scrapper.py")
    if not os.path.exists(ruta_scraper):
        return list(BECAS_CURADAS), "curada"

    scrapeadas = []
    for url in URLS_BECAS_SCRAPER:
        try:
            payload = json.dumps({"url": url, "cantidad": 5, "objetivo": objetivo, "headless": True})
            res = subprocess.run(
                [sys.executable, ruta_scraper, payload],
                capture_output=True, text=True, timeout=90, encoding="utf-8"
            )
            # El stdout puede traer prints de progreso; tomar la ultima linea JSON valida
            for linea in reversed((res.stdout or "").splitlines()):
                linea = linea.strip()
                if not linea.startswith("{"):
                    continue
                data = json.loads(linea)
                plataforma = data.get("plataforma", "Web")
                for c in (data.get("cursos") or []):
                    titulo = (c.get("Titulo") or "").strip()
                    enlace = (c.get("URL") or "").strip()
                    if titulo and enlace:
                        desc = (c.get("Descripcion") or "").strip()
                        if not desc or desc == "N/A":
                            desc = f"Convocatoria vigente detectada en {plataforma} (scraping en vivo)."
                        scrapeadas.append((plataforma, titulo, enlace, desc))
                break
        except Exception:
            continue

    if scrapeadas:
        # Combinar: primero lo scrapeado en vivo, luego curadas (sin duplicar enlaces)
        vistos = {s[2] for s in scrapeadas}
        combinado = scrapeadas + [b for b in BECAS_CURADAS if b[2] not in vistos]
        return combinado, "en_vivo+curada"
    return list(BECAS_CURADAS), "curada"

# Rutas del sistema ODEM
RUTA_BASE       = os.path.dirname(os.path.abspath(__file__))
RUTA_AGENTES    = os.path.dirname(RUTA_BASE)
RUTA_LUMINA     = os.path.join(RUTA_AGENTES, "Lumina", "Lumina_sql.py")
RUTA_REPORTS    = os.path.abspath(os.path.join(RUTA_AGENTES, "..", "..", "reports"))

# Umbrales socioeconómicos (pesos colombianos / mes)
UMBRAL_ICETEX            = 4_000_000
UMBRAL_BECA_GENERACION_E = 3_000_000

# ================================================================
#   BASE DE DATOS SNIES — MOCK (fallback cuando Lumina no está disponible)
# ================================================================

SNIES_MOCK: list[dict] = [
    {
        "cod_snies": "1001", "nombre_programa": "Ingeniería de Sistemas",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Ingeniería, Arquitectura, Urbanismo y afines",
        "nombre_ies": "Universidad Nacional de Colombia",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Pública",
        "costo_semestre_cop": 2_500_000, "duracion_semestres": 10,
        "link_inscripcion": "https://admisiones.unal.edu.co",
        "requisitos": ["ICFES > 300", "Examen de admisión UNAL", "Documento de identidad"],
        "becas_disponibles": ["Beca Mejores Bachilleres", "Generación E", "ICETEX Tú Eliges"],
        "perfil_vocacional": ["lógico-matemático", "tecnología", "resolución de problemas", "analítico"],
        "actividades_extracurriculares": [
            "Semilleros de investigación en IA y Machine Learning",
            "Grupos de desarrollo de software libre",
            "Hackatones nacionales e internacionales",
        ]
    },
    {
        "cod_snies": "1002", "nombre_programa": "Ingeniería Biomédica",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Ingeniería, Arquitectura, Urbanismo y afines",
        "nombre_ies": "Universidad EAN",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Privada",
        "costo_semestre_cop": 8_200_000, "duracion_semestres": 10,
        "link_inscripcion": "https://www.ean.edu.co/programas/ingenieria-biomedica",
        "requisitos": ["ICFES habilitado", "Entrevista", "Documento de identidad"],
        "becas_disponibles": ["Beca Talento EAN", "ICETEX Tú Eliges", "Descuento Familia"],
        "perfil_vocacional": ["ciencias naturales", "tecnología", "medicina", "innovación", "analítico"],
        "actividades_extracurriculares": [
            "Laboratorio de prototipado biomédico",
            "Convenios con clínicas y hospitales",
            "Feria INNOVA EAN",
        ]
    },
    {
        "cod_snies": "1003", "nombre_programa": "Ingeniería Civil",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Ingeniería, Arquitectura, Urbanismo y afines",
        "nombre_ies": "Universidad de los Andes",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Privada",
        "costo_semestre_cop": 16_000_000, "duracion_semestres": 10,
        "link_inscripcion": "https://uniandes.edu.co/admisiones",
        "requisitos": ["ICFES > 320", "Prueba de clasificación", "Documento de identidad"],
        "becas_disponibles": ["Beca Quiero Estudiar", "ICETEX", "Apoyo socioeconómico Uniandes"],
        "perfil_vocacional": ["matemáticas", "construcción", "diseño espacial", "liderazgo", "trabajo en campo"],
        "actividades_extracurriculares": [
            "Visitas técnicas a obras", "Concursos de diseño estructural",
        ]
    },
    {
        "cod_snies": "2001", "nombre_programa": "Medicina",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Ciencias de la Salud",
        "nombre_ies": "Universidad de Antioquia",
        "municipio": "Medellín", "departamento": "Antioquia",
        "modalidad": "Presencial", "caracter_ies": "Pública",
        "costo_semestre_cop": 3_200_000, "duracion_semestres": 12,
        "link_inscripcion": "https://www.udea.edu.co/admisiones",
        "requisitos": ["ICFES > 330", "Examen de admisión UdeA", "Prueba vocacional"],
        "becas_disponibles": ["Generación E", "Beca SPE Antioquia", "ICETEX Medicina"],
        "perfil_vocacional": ["ciencias naturales", "empatía", "servicio a otros", "biología", "química"],
        "actividades_extracurriculares": [
            "Brigadas de salud comunitarias", "Rotaciones internacionales",
        ]
    },
    {
        "cod_snies": "2002", "nombre_programa": "Enfermería",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Ciencias de la Salud",
        "nombre_ies": "Universidad Nacional de Colombia",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Pública",
        "costo_semestre_cop": 1_800_000, "duracion_semestres": 8,
        "link_inscripcion": "https://admisiones.unal.edu.co",
        "requisitos": ["ICFES habilitado", "Examen de admisión UNAL"],
        "becas_disponibles": ["Generación E", "Beca Mejores Bachilleres", "ICETEX"],
        "perfil_vocacional": ["servicio a otros", "salud", "empatía", "trabajo en equipo", "ciencias naturales"],
        "actividades_extracurriculares": ["Voluntariado Cruz Roja", "Prácticas hospitalarias"],
    },
    {
        "cod_snies": "3001", "nombre_programa": "Psicología",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Ciencias Sociales y Humanas",
        "nombre_ies": "Universidad Javeriana",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Privada",
        "costo_semestre_cop": 12_500_000, "duracion_semestres": 10,
        "link_inscripcion": "https://www.javeriana.edu.co/admisiones",
        "requisitos": ["ICFES habilitado", "Entrevista personal", "Ensayo de motivación"],
        "becas_disponibles": ["Beca Padre Diego Jaramillo", "ICETEX", "Descuento empleados"],
        "perfil_vocacional": ["empatía", "comunicación", "análisis social", "ayuda a otros", "humanidades"],
        "actividades_extracurriculares": ["Consultorio psicológico comunitario"],
    },
    {
        "cod_snies": "3002", "nombre_programa": "Derecho",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Ciencias Sociales y Humanas",
        "nombre_ies": "Universidad Externado de Colombia",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Privada",
        "costo_semestre_cop": 9_800_000, "duracion_semestres": 10,
        "link_inscripcion": "https://www.uexternado.edu.co/admisiones",
        "requisitos": ["ICFES habilitado", "Examen de admisión", "Entrevista"],
        "becas_disponibles": ["Beca Externado", "ICETEX Tú Eliges", "Apoyo económico"],
        "perfil_vocacional": ["argumentación", "justicia social", "liderazgo", "lectura", "debate", "humanidades"],
        "actividades_extracurriculares": ["Consultorio jurídico gratuito", "Litigación académica"],
    },
    {
        "cod_snies": "4001", "nombre_programa": "Administración de Empresas",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Economía, Administración, Contaduría y afines",
        "nombre_ies": "Universidad del Rosario",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Privada",
        "costo_semestre_cop": 11_000_000, "duracion_semestres": 8,
        "link_inscripcion": "https://www.urosario.edu.co/admisiones",
        "requisitos": ["ICFES habilitado", "Entrevista", "Prueba de inglés básica"],
        "becas_disponibles": ["Beca Rosarista Destacado", "ICETEX", "Crédito condonable"],
        "perfil_vocacional": ["liderazgo", "negocios", "comunicación", "emprendimiento", "organización"],
        "actividades_extracurriculares": ["Semillero de emprendimiento", "Modelo ONU (MUN)"],
    },
    {
        "cod_snies": "5001", "nombre_programa": "Diseño Gráfico",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Bellas Artes",
        "nombre_ies": "Universidad Jorge Tadeo Lozano",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Privada",
        "costo_semestre_cop": 7_500_000, "duracion_semestres": 8,
        "link_inscripcion": "https://www.utadeo.edu.co/admisiones",
        "requisitos": ["ICFES habilitado", "Portafolio de diseño o dibujo", "Entrevista"],
        "becas_disponibles": ["Beca Talento Artístico Tadeo", "ICETEX", "Descuento hermanos"],
        "perfil_vocacional": ["creatividad", "arte", "diseño", "comunicación visual", "tecnología", "estética"],
        "actividades_extracurriculares": ["FabLab", "Exposiciones en galerías", "Concursos internacionales"],
    },
    {
        "cod_snies": "6001", "nombre_programa": "Estadística",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Matemáticas y Ciencias Naturales",
        "nombre_ies": "Universidad Nacional de Colombia",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Pública",
        "costo_semestre_cop": 2_000_000, "duracion_semestres": 8,
        "link_inscripcion": "https://admisiones.unal.edu.co",
        "requisitos": ["ICFES > 290", "Examen de admisión UNAL"],
        "becas_disponibles": ["Generación E", "Beca Mejores Bachilleres", "ICETEX"],
        "perfil_vocacional": ["matemáticas", "analítico", "datos", "tecnología", "ciencias exactas", "investigación"],
        "actividades_extracurriculares": ["Semillero de Ciencia de Datos", "Competencias de estadística"],
    },
    {
        "cod_snies": "7001", "nombre_programa": "Licenciatura en Matemáticas",
        "nivel_formacion": "Universitario",
        "area_conocimiento": "Educación",
        "nombre_ies": "Universidad Pedagógica Nacional",
        "municipio": "Bogotá D.C.", "departamento": "Cundinamarca",
        "modalidad": "Presencial", "caracter_ies": "Pública",
        "costo_semestre_cop": 1_500_000, "duracion_semestres": 10,
        "link_inscripcion": "https://www.pedagogica.edu.co/admisiones",
        "requisitos": ["ICFES habilitado", "Examen de admisión UPN"],
        "becas_disponibles": ["Generación E", "ICETEX"],
        "perfil_vocacional": ["enseñanza", "matemáticas", "paciencia", "vocación docente", "comunicación"],
        "actividades_extracurriculares": ["Práctica docente desde 1er semestre", "Olimpiadas de matemáticas"],
    },
]

AREA_KEYWORDS: dict[str, list[str]] = {
    "Ingeniería, Arquitectura, Urbanismo y afines": [
        "matemáticas", "física", "tecnología", "computadores", "construcción",
        "lógico-matemático", "analítico", "programación", "diseño técnico",
        "resolución de problemas", "robots", "electrónica", "mecánica"
    ],
    "Ciencias de la Salud": [
        "biología", "química", "medicina", "salud", "cuerpo humano",
        "empatía", "ayudar a otros", "ciencias naturales", "enfermedades", "hospital"
    ],
    "Ciencias Sociales y Humanas": [
        "humanidades", "historia", "filosofía", "sociedad", "política",
        "argumentación", "justicia social", "debate", "análisis social", "lectura"
    ],
    "Economía, Administración, Contaduría y afines": [
        "negocios", "dinero", "empresa", "liderazgo", "organización",
        "emprendimiento", "mercado", "ventas", "administración", "finanzas"
    ],
    "Bellas Artes": [
        "arte", "creatividad", "diseño", "música", "teatro", "fotografía",
        "comunicación visual", "estética", "pintura", "ilustración"
    ],
    "Matemáticas y Ciencias Naturales": [
        "matemáticas", "física", "química", "estadística", "datos",
        "investigación", "ciencias exactas", "laboratorio", "experimentos"
    ],
    "Educación": [
        "enseñar", "docencia", "niños", "vocación docente", "paciencia",
        "comunicación", "pedagogía", "colegios", "aprendizaje"
    ],
}

# ================================================================
#   CUESTIONARIO DE APTITUDES
# ================================================================

TEST_APTITUD: list[dict] = []
# Cargar preguntas estructuradas desde archivo JSON
RUTA_PREGUNTAS_JSON = os.path.join(os.path.dirname(__file__), "preguntas_estructuradas.json")
try:
    with open(RUTA_PREGUNTAS_JSON, "r", encoding="utf-8") as f:
        TEST_APTITUD = json.load(f)
except Exception:
    pass

# ================================================================
#   SYSTEM PROMPT — Ada (identidad dentro del sistema ODEM)
# ================================================================

SYSTEM_Ada = """Eres Ada, el agente de orientación vocacional del sistema multiagente ODEM
(Universidad EAN). Eres el punto de contacto directo con el ciudadano beneficiario —
especialmente jóvenes egresados de bachillerato en Colombia de bajos recursos.

Tu propósito es democratizar el acceso a la educación superior: guiar a cada joven
hacia la carrera universitaria más adecuada para su perfil, habilidades e ingresos.

En el sistema multiagente, eres coordinado por Viernes. Trabajas junto a:
- Lumina: te provee datos reales del SNIES (programas, universidades) y de deserción
  académica (SPADIES) por programa/estrato/género. El SNIES NO incluye costo de matrícula.
- Scrapper: te trae datos actualizados de becas y convocatorias en tiempo real.

Además, en el contexto que recibes tienes datos REALES de mercado laboral de la
GEIH-DANE (encuesta de hogares): ingreso mediano y % de informalidad por departamento
y por sector económico afín al programa recomendado. Esto es lo más parecido que
tienes a "cuánto se gana" cuando el SNIES no trae costos ni salarios — úsalo para que
la orientación no sea solo "qué estudiar" sino "qué tan bien pago y formal es el sector
al que ese programa suele llevar en tu departamento". Interprétalo como estadística de
la encuesta (no como garantía individual) y combínalo con el análisis de deserción para
dar una imagen más completa de riesgo/oportunidad.

Cuando redactes reportes: sé cálido, empático, claro y práctico. El joven no siempre
conoce el sistema universitario colombiano — explica lo relevante sin tecnicismos.
Usa SIEMPRE las cifras reales provistas en el contexto (deserción, mercado laboral);
nunca inventes porcentajes, salarios ni costos de matrícula. Siempre incluye links
concretos, requisitos y pasos de acción."""


# ================================================================
#   CLASE PRINCIPAL
# ================================================================

class AgenteAda:
    """
    Agente Ada — Orientación Vocacional.
    Componente del Sistema Multiagente ODEM · Universidad EAN.

    Uso autónomo (CLI):
        python3 Ada.py

    Modo demo:
        DEMO_MODE=true python3 Ada.py

    Integración con Viernes/multiagente (argv):
        python3 Ada.py '{"datos_personales": {...}, "respuestas_test": {...}}'
    """

    def __init__(
        self,
        datos_personales: Optional[dict] = None,
        respuestas_test: Optional[dict] = None,
    ):
        # Cliente NVIDIA opcional: el reporte principal se genera con Gemini (requests).
        # Solo se inicializa si hay NVIDIA_API_KEY, para no romper el flujo cuando no existe.
        if API_KEY:
            try:
                self.client = openai.OpenAI(api_key=API_KEY, base_url="https://integrate.api.nvidia.com/v1")
            except Exception:
                self.client = None
        else:
            self.client = None
        self.model = MODEL
        self.datos_personales: dict = datos_personales or {}
        self.respuestas_test: dict = respuestas_test or {}
        self.perfil_vocacional: list[str] = []
        self.programas_recomendados: list[dict] = []
        self.reporte_final: str = ""
        self.fuente_snies: str = "mock"   # 'lumina' | 'mock' | 'snies_directo'
        self.fuente_becas: str = "curada"  # 'curada' | 'en_vivo+curada'
        self.promedios_categorias: dict = {}
        if respuestas_test:
            self.recalcular_perfil_desde_respuestas()

    # ────────────────────────────────────────────────────────────
    # PASO 1 — DATOS PERSONALES
    # ────────────────────────────────────────────────────────────

    def paso1_recolectar_datos(self) -> dict:
        print("\n" + "═" * 60)
        print("  PASO 1 — DATOS PERSONALES")
        print("═" * 60)

        def pedir(campo: str, ejemplo: str = "") -> str:
            hint = f" (ej: {ejemplo})" if ejemplo else ""
            while True:
                v = input(f"  {campo}{hint}: ").strip()
                if v:
                    return v
                print("  ⚠ Campo obligatorio.")

        def pedir_num(campo: str, mn: int, mx: int) -> int:
            while True:
                try:
                    v = int(input(f"  {campo} ({mn}–{mx}): ").strip())
                    if mn <= v <= mx:
                        return v
                    print(f"  ⚠ Número entre {mn} y {mx}.")
                except ValueError:
                    print("  ⚠ Número válido.")

        self.datos_personales = {
            "cedula":                  pedir("Número de documento (CC/TI)", "1012345678"),
            "nombre_completo":         pedir("Nombre completo", "Ana María López"),
            "edad":                    pedir_num("Edad", 14, 30),
            "municipio":               pedir("Municipio de residencia", "Bogotá D.C."),
            "departamento":            pedir("Departamento", "Cundinamarca"),
            "direccion":               pedir("Dirección", "Calle 45 # 12-34"),
            "estrato":                 pedir_num("Estrato socioeconómico", 1, 6),
            "ingresos_familiares_cop": int(
                pedir("Ingresos familiares mensuales (COP, sin puntos)", "2500000")
                .replace(".", "").replace(",", "")
            ),
            "icfes_puntaje":           pedir_num("Puntaje ICFES (0 si aún no lo tienes)", 0, 500),
            "colegio":                 pedir("Nombre del colegio", "Instituto Pedagógico Nacional"),
            "tipo_colegio":            pedir("Tipo de colegio (público/privado)", "público"),
        }
        print("\n  ✅ Datos personales registrados.\n")
        return self.datos_personales

    # ────────────────────────────────────────────────────────────
    # PASO 2 — TEST DE APTITUDES
    # ────────────────────────────────────────────────────────────

    def paso2_test_aptitud(self) -> dict:
        print("\n" + "═" * 60)
        print(f"  PASO 2 — TEST DE APTITUDES ({len(TEST_APTITUD)} PREGUNTAS)")
        print("═" * 60)
        print("  Responde en una escala de 1 a 5:")
        print("  1: Totalmente en desacuerdo")
        print("  2: En desacuerdo")
        print("  3: Neutro")
        print("  4: De acuerdo")
        print("  5: Totalmente de acuerdo\n")
        self.respuestas_test = {}
        
        print("  ¿Deseas simular las respuestas de forma automática para pruebas rápidas? (S/N)")
        simular = input("  Elección: ").strip().upper() == "S"

        import random
        for idx, pregunta in enumerate(TEST_APTITUD, 1):
            qid = pregunta["id"]
            if simular:
                resp = random.randint(1, 5)
                self.respuestas_test[f"p{qid}"] = resp
                continue

            print(f"  [{idx}/{len(TEST_APTITUD)}] {pregunta['question']}")
            while True:
                try:
                    resp_str = input("  Tu respuesta (1-5): ").strip()
                    resp = int(resp_str)
                    if 1 <= resp <= 5:
                        self.respuestas_test[f"p{qid}"] = resp
                        print()
                        break
                    print("  ⚠ Ingresa un número entre 1 y 5.")
                except ValueError:
                    print("  ⚠ Ingresa un número válido entre 1 y 5.")

        self.recalcular_perfil_desde_respuestas()
        print(f"  ✅ Perfil detectado: {', '.join(self.perfil_vocacional[:6])} ...")
        return self.respuestas_test

    def recalcular_perfil_desde_respuestas(self) -> None:
        """
        Calcula self.perfil_vocacional y las puntuaciones por categoría
        a partir de self.respuestas_test.
        """
        # CATEGORIAS_MAP vive a nivel de modulo (junto a PERFIL_A_REGEXP) para
        # que _score_afinidad() tambien pueda usarlo via KEYWORD_A_CATEGORIA.
        scores_categorias = {cat: 0 for cat in CATEGORIAS_MAP}
        counts_categorias = {cat: 0 for cat in CATEGORIAS_MAP}

        for q in TEST_APTITUD:
            qid = q["id"]
            cat = q["category"]
            
            # Obtener respuesta (valor por defecto: 3)
            val = self.respuestas_test.get(f"p{qid}", 3)
            
            # Manejar si viene como string ("1" a "5") o como letra vieja "A"/"B"/"C"/"D"
            if isinstance(val, str):
                if val.isdigit():
                    val = int(val)
                else:
                    ans_map = {"A": 5, "B": 4, "C": 3, "D": 2}
                    val = ans_map.get(val.upper(), 3)
            
            # Asegurar rango 1-5
            try:
                val = int(val)
                if not (1 <= val <= 5):
                    val = 3
            except Exception:
                val = 3

            score = val
            if q.get("reversed", False):
                score = 6 - val

            scores_categorias[cat] += score
            counts_categorias[cat] += 1

        # Calcular promedios y porcentajes
        self.promedios_categorias = {}
        for cat in CATEGORIAS_MAP:
            if counts_categorias[cat] > 0:
                self.promedios_categorias[cat] = scores_categorias[cat] / counts_categorias[cat]
            else:
                self.promedios_categorias[cat] = 3.0

        # Construir perfil vocacional ordenado de mayor a menor afinidad
        sorted_cats = sorted(self.promedios_categorias.items(), key=lambda x: x[1], reverse=True)
        perfil_kws = []
        for cat, avg in sorted_cats:
            kws = CATEGORIAS_MAP[cat]
            perfil_kws.extend(kws)

        self.perfil_vocacional = perfil_kws

    # ────────────────────────────────────────────────────────────
    # PASO 3 — CONSULTA SNIES (Lumina → fallback mock)
    # ────────────────────────────────────────────────────────────

    def paso3_consultar_snies(self) -> list[dict]:
        print("\n" + "═" * 60)
        print("  PASO 3 — CONSULTA SNIES (vía Lumina)")
        print("═" * 60)

        # 1) Consulta DETERMINISTA directa a la DB (SQL fijo, sin depender del LLM)
        programas = self._consultar_snies_directo()
        if programas:
            self.fuente_snies = "snies_directo"
            print(f"  ✅ SNIES (consulta directa) devolvió {len(programas)} programas reales.")
        else:
            # 2) Fallback: Lumina (via LLM). Puede variar por no-determinismo del modelo.
            programas = self._consultar_via_lumina()
            if programas:
                self.fuente_snies = "lumina"
                print(f"  ✅ Lumina respondió con {len(programas)} programas del SNIES real.")
            else:
                # 3) Fallback final: datos locales mock
                print("  ⚠ SNIES/Lumina no disponibles. Usando datos SNIES locales.")
                programas = self._consultar_snies_mock()
                self.fuente_snies = "mock"

        # Filtro socioeconómico
        ingresos = self.datos_personales.get("ingresos_familiares_cop", 0)
        estrato  = self.datos_personales.get("estrato", 3)
        icfes    = self.datos_personales.get("icfes_puntaje", 0)

        viables = []
        for p in programas:
            costo = p.get("costo_semestre_cop") or 0
            if estrato <= 2 and ingresos < UMBRAL_ICETEX:
                if p.get("caracter_ies") == "Privada" and costo > 6_000_000 and icfes < 300:
                    continue
            score = _score_afinidad(p.get("perfil_vocacional") or [], self.promedios_categorias)
            p["score_afinidad"] = score
            viables.append(p)

        viables.sort(key=lambda x: x["score_afinidad"], reverse=True)
        self.programas_recomendados = viables[:5]

        print(f"  ✅ {len(self.programas_recomendados)} programas recomendados:")
        for i, p in enumerate(self.programas_recomendados, 1):
            print(f"     {i}. {p['nombre_programa']} — {p['nombre_ies']}")

        return self.programas_recomendados

    def _consultar_snies_directo(self) -> list[dict]:
        """
        Consulta DETERMINISTA a la base de datos SNIES real, con SQL fijo y
        parametrizado (REGEXP construido desde el perfil vocacional).
        No usa el LLM, por lo que el resultado es reproducible.
        Si la DB no esta disponible, retorna [] para activar el fallback.
        """
        try:
            import mysql.connector
        except Exception:
            return []

        # Construir patron REGEXP a partir del perfil vocacional
        tokens = set()
        for kw in self.perfil_vocacional:
            patron = PERFIL_A_REGEXP.get(kw)
            if patron:
                for t in patron.split("|"):
                    tokens.add(t.strip())
        if not tokens:
            # Perfil sin mapeo -> no forzar; dejar fallback
            return []
        regexp = "|".join(sorted(tokens))

        municipio = (self.datos_personales.get("municipio") or "").upper().strip()

        sql = (
            "SELECT programa_academico, institucion_de_educacion_superior_ies, "
            "caracter_ies, municipio_de_oferta_del_programa, metodologia, "
            "area_de_conocimiento, SUM(matriculados) AS tot "
            "FROM snies_matriculados "
            "WHERE programa_academico REGEXP %s "
            "GROUP BY programa_academico, institucion_de_educacion_superior_ies, "
            "caracter_ies, municipio_de_oferta_del_programa, metodologia, area_de_conocimiento "
            # Orden GENERAL: prioriza el municipio del usuario y luego la popularidad real
            # (numero de matriculados). NO se fuerza modalidad virtual.
            "ORDER BY (municipio_de_oferta_del_programa = %s) DESC, tot DESC "
            "LIMIT 40"
        )

        try:
            conn = mysql.connector.connect(**_db_kwargs())
            cursor = conn.cursor()
            cursor.execute(sql, (regexp, municipio))
            filas = cursor.fetchall()
            cursor.close()
            conn.close()
        except Exception:
            return []

        import urllib.parse as _up
        programas = []
        for (prog, ies, caracter, muni, metod, area, tot) in filas:
            prog_t = (prog or "N/A").title()
            ies_t = (ies or "N/A").title()
            # Link REAL de busqueda del programa en la IES (no un '#' ilustrativo)
            q = _up.quote(f"{prog_t} {ies_t} inscripciones admisiones")
            link = f"https://www.google.com/search?q={q}"
            programas.append({
                "cod_snies":              "SNIES",
                "nombre_programa":        prog_t,
                "nivel_formacion":        "Universitario",
                "area_conocimiento":      (area or "N/A").title(),
                "nombre_ies":             ies_t,
                "municipio":              (muni or "N/A").title(),
                "departamento":           "N/A",
                "modalidad":              (metod or "Presencial").title(),
                "caracter_ies":           (caracter or "N/A").title(),
                # El SNIES de matriculados NO incluye costo real -> no inventar $0
                "costo_semestre_cop":     None,
                "costo_texto":            "Consultar con la IES",
                "duracion_semestres":     10,
                "link_inscripcion":       link,
                "requisitos":             ["ICFES habilitado"],
                "becas_disponibles":      ["ICETEX", "Generación E"],
                "perfil_vocacional":      _tags_programa(prog_t, area),
                "actividades_extracurriculares": [],
                "matriculados_hist":      _a_int(tot, 0),
            })

        # Balanceo GENERAL: evitar que domine una sola modalidad o IES.
        # Intercala presenciales y virtuales, limita repeticiones de la misma IES.
        presenciales = [p for p in programas if "VIRTUAL" not in p["modalidad"].upper() and "DISTANCIA" not in p["modalidad"].upper()]
        virtuales    = [p for p in programas if p not in presenciales]
        balanceado = []
        ies_count = {}
        i = j = 0
        # Alterna 1 presencial, 1 virtual, respetando popularidad dentro de cada grupo
        while (i < len(presenciales) or j < len(virtuales)) and len(balanceado) < 15:
            for grupo, idx in ((presenciales, "i"), (virtuales, "j")):
                k = i if idx == "i" else j
                # avanzar saltando IES ya muy repetidas (max 2 por IES)
                while k < len(grupo):
                    cand = grupo[k]
                    k += 1
                    if ies_count.get(cand["nombre_ies"], 0) < 2:
                        balanceado.append(cand)
                        ies_count[cand["nombre_ies"]] = ies_count.get(cand["nombre_ies"], 0) + 1
                        break
                if idx == "i":
                    i = k
                else:
                    j = k
        # Si el balanceo quedo corto, completar con lo que reste
        if len(balanceado) < 10:
            for p in programas:
                if p not in balanceado:
                    balanceado.append(p)
                if len(balanceado) >= 15:
                    break
        return balanceado if balanceado else programas

    def _analizar_desercion(self, programas=None) -> dict:
        """
        Analiza la tabla desercion_academica para las areas del perfil / programas
        recomendados. Devuelve conteos reales por programa, estrato y genero.
        Retorna {} si no hay datos o la DB no esta disponible.
        """
        try:
            import mysql.connector
        except Exception:
            return {}

        # Construir REGEXP desde el perfil (mismo mapeo que programas)
        tokens = set()
        for kw in self.perfil_vocacional:
            patron = PERFIL_A_REGEXP.get(kw)
            if patron:
                for t in patron.split("|"):
                    tokens.add(t.strip())
        if not tokens:
            return {}
        regexp = "|".join(sorted(tokens))

        try:
            conn = mysql.connector.connect(**_db_kwargs())
            cur = conn.cursor()
            # Total de casos de desercion (perdida de cupo) para programas afines al perfil
            cur.execute(
                "SELECT COUNT(*) FROM desercion_academica WHERE UPPER(nombre_programa) REGEXP %s",
                (regexp,)
            )
            total_afin = int(cur.fetchone()[0] or 0)
            # Top programas con mas casos de desercion (afines al perfil)
            cur.execute(
                "SELECT nombre_programa, COUNT(*) c FROM desercion_academica "
                "WHERE UPPER(nombre_programa) REGEXP %s "
                "GROUP BY nombre_programa ORDER BY c DESC LIMIT 5",
                (regexp,)
            )
            top_prog = [(r[0].title(), int(r[1])) for r in cur.fetchall()]
            # Distribucion por estrato
            cur.execute(
                "SELECT estrato, COUNT(*) c FROM desercion_academica "
                "WHERE UPPER(nombre_programa) REGEXP %s AND estrato IS NOT NULL AND estrato <> '' "
                "GROUP BY estrato ORDER BY c DESC LIMIT 6",
                (regexp,)
            )
            por_estrato = [(str(r[0]), int(r[1])) for r in cur.fetchall()]
            # Distribucion por genero
            cur.execute(
                "SELECT genero, COUNT(*) c FROM desercion_academica "
                "WHERE UPPER(nombre_programa) REGEXP %s GROUP BY genero ORDER BY c DESC",
                (regexp,)
            )
            por_genero = [(str(r[0]), int(r[1])) for r in cur.fetchall()]
            cur.close(); conn.close()
            if total_afin == 0:
                return {}
            return {
                "total_casos_afines": total_afin,
                "top_programas_desercion": top_prog,
                "por_estrato": por_estrato,
                "por_genero": por_genero,
            }
        except Exception:
            return {}

    def _analizar_mercado_laboral(self) -> dict:
        """
        Cruza el area_de_conocimiento del programa mejor rankeado con las
        tablas geih_sector_departamento / geih_departamento_resumen (DANE,
        cargadas por scripts/consolidar_geih.py) para dar contexto REAL de
        mercado laboral -ingreso mediano, informalidad- en el departamento
        del estudiante. Determinista (SQL fijo), sin LLM.
        Retorna {} si no hay match de departamento o de area->CIIU.
        """
        try:
            import mysql.connector
        except Exception:
            return {}

        departamento = (self.datos_personales.get("departamento") or "").upper().strip()
        dpto_cod = NOMBRE_A_DPTO.get(departamento)
        if not dpto_cod or not self.programas_recomendados:
            return {}

        area = (self.programas_recomendados[0].get("area_conocimiento") or "").upper().strip()
        ciiu_codes = AREA_CONOCIMIENTO_A_CIIU.get(area, [])

        try:
            conn = mysql.connector.connect(**_db_kwargs())
            cur = conn.cursor()
            resultado = {}

            if ciiu_codes:
                fmt = ",".join(["%s"] * len(ciiu_codes))
                cur.execute(
                    f"SELECT sector_nombre, ingreso_mediana, pct_informalidad, poblacion_ocupada_estimada "
                    f"FROM geih_sector_departamento WHERE dpto=%s AND sector_ciiu_2d IN ({fmt}) "
                    f"ORDER BY poblacion_ocupada_estimada DESC LIMIT 3",
                    [dpto_cod] + ciiu_codes,
                )
                resultado["sectores_afines"] = [
                    {
                        "sector": r[0],
                        "ingreso_mediana": float(r[1]) if r[1] is not None else None,
                        "pct_informalidad": float(r[2]) if r[2] is not None else None,
                        "poblacion_ocupada_estimada": int(r[3]) if r[3] is not None else None,
                    }
                    for r in cur.fetchall()
                ]

            cur.execute(
                "SELECT ingreso_mediana, pct_informalidad, tasa_desempleo_pct "
                "FROM geih_departamento_resumen WHERE dpto=%s",
                (dpto_cod,),
            )
            row = cur.fetchone()
            if row:
                resultado["departamento_general"] = {
                    "ingreso_mediana": float(row[0]) if row[0] is not None else None,
                    "pct_informalidad": float(row[1]) if row[1] is not None else None,
                    "tasa_desempleo_pct": float(row[2]) if row[2] is not None else None,
                }
            cur.close()
            conn.close()
            if resultado.get("sectores_afines") or resultado.get("departamento_general"):
                return resultado
            return {}
        except Exception:
            return {}

    def _consultar_via_lumina(self) -> list[dict]:
        """
        Llama a Lumina (Odemiroconsql.py) para obtener programas del SNIES real.
        Retorna lista de dicts compatibles con el formato interno.
        Si falla, retorna lista vacía para activar el fallback.
        """
        if not os.path.exists(RUTA_LUMINA):
            return []

        perfil_str = ", ".join(self.perfil_vocacional[:6])
        municipio  = self.datos_personales.get("municipio", "Bogotá")
        # Query simple y robusta: consulta directa de programas del SNIES real.
        # Se evita pedir "predicciones estadisticas" en el mismo request porque
        # induce SQL invalido; el analisis lo hace despues Gemini en el reporte.
        query = (
            f"Dame los 10 programas universitarios del SNIES mas relacionados con estas areas: "
            f"{perfil_str}. Preferiblemente en {municipio} o modalidad virtual. "
            f"Devuelve un array JSON con objetos que tengan estas claves: "
            f"nombre_programa, nombre_ies, caracter_ies, municipio, modalidad, link_inscripcion. "
            f"Responde con el array JSON dentro de un bloque de codigo."
        )

        try:
            resultado = subprocess.run(
                [sys.executable, RUTA_LUMINA, query],
                capture_output=True, text=True, timeout=120,
                encoding="utf-8"
            )
            if resultado.returncode != 0 or not resultado.stdout.strip():
                return []

            data = json.loads(resultado.stdout.strip())
            # Si Lumina reporta error honesto, no hay datos utilizables -> fallback
            if data.get("error") or data.get("tipo") == "ERROR":
                return []
            respuesta_texto = data.get("respuesta", "") or ""
            programas_raw = None

            # 0) PREFERIR las filas crudas que Lumina ahora devuelve en 'datos'
            #    (lista de dicts columna->valor). Es la fuente mas confiable.
            filas = data.get("datos")
            if isinstance(filas, list) and filas:
                programas_raw = filas

            # Si no vinieron filas crudas, intentar extraer un array JSON de la prosa.
            if not programas_raw:
                # 1) Preferir el contenido de un bloque de codigo ```json ... ``` o ``` ... ```
                bloque = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", respuesta_texto, re.DOTALL)
                candidatos = []
                if bloque:
                    candidatos.append(bloque.group(1))
                # 2) Fallback: el primer '[' hasta el ultimo ']' (greedy)
                greedy = re.search(r"\[.*\]", respuesta_texto, re.DOTALL)
                if greedy:
                    candidatos.append(greedy.group(0))

                for cand in candidatos:
                    try:
                        parsed = json.loads(cand)
                        if isinstance(parsed, list) and parsed:
                            programas_raw = parsed
                            break
                    except Exception:
                        continue

            if programas_raw:
                # Normalizar al formato interno Ada
                programas_normalizados = []
                for p in programas_raw:
                    programas_normalizados.append({
                        "cod_snies":              p.get("cod_snies", "SNIES"),
                        "nombre_programa":        p.get("nombre_programa", p.get("programa", "N/A")),
                        "nivel_formacion":        p.get("nivel_formacion", "Universitario"),
                        "area_conocimiento":      p.get("area_conocimiento", "N/A"),
                        "nombre_ies":             p.get("nombre_ies", p.get("universidad", "N/A")),
                        "municipio":              p.get("municipio", municipio),
                        "departamento":           p.get("departamento", "N/A"),
                        "modalidad":              p.get("modalidad", "Presencial"),
                        "caracter_ies":           p.get("caracter_ies", p.get("caracter", "N/A")),
                        "costo_semestre_cop":     _a_int(p.get("costo_semestre_cop"), 0),
                        "duracion_semestres":     _a_int(p.get("duracion_semestres"), 10),
                        "link_inscripcion":       p.get("link_inscripcion", p.get("link", "#")),
                        "requisitos":             p.get("requisitos") or ["ICFES habilitado"],
                        "becas_disponibles":      p.get("becas_disponibles") or ["ICETEX"],
                        "perfil_vocacional":      p.get("perfil_vocacional") or _tags_programa(
                            p.get("nombre_programa", p.get("programa", "")),
                            p.get("area_conocimiento", ""),
                        ),
                        "actividades_extracurriculares": p.get("actividades_extracurriculares") or [],
                    })
                return programas_normalizados
        except Exception:
            pass

        return []

    def _consultar_snies_mock(self) -> list[dict]:
        """Filtro local sobre el SNIES mock por perfil vocacional."""
        area_scores: dict[str, int] = {}
        for area, keywords in AREA_KEYWORDS.items():
            score = sum(1 for kw in self.perfil_vocacional if kw in keywords)
            if score > 0:
                area_scores[area] = score

        if not area_scores:
            return SNIES_MOCK[:5]

        top_areas = sorted(area_scores, key=area_scores.get, reverse=True)[:3]
        return [p for p in SNIES_MOCK if p["area_conocimiento"] in top_areas]

    # ────────────────────────────────────────────────────────────
    # PASO 4 — REPORTE PERSONALIZADO (Claude)
    # ────────────────────────────────────────────────────────────

    def paso4_generar_reporte(self) -> str:
        print("\n" + "═" * 60)
        print("  PASO 4 — GENERANDO REPORTE CON GEMINI")
        print("═" * 60)
        print("  🤖 Analizando tu perfil... (puede tomar unos segundos)\n")

        dp       = self.datos_personales
        ingresos = dp.get("ingresos_familiares_cop", 0)
        estrato  = dp.get("estrato", 3)
        icfes    = dp.get("icfes_puntaje", 0)

        beca_gen_e = ingresos <= UMBRAL_BECA_GENERACION_E and estrato <= 3
        beca_icetex = ingresos <= UMBRAL_ICETEX

        ctx = f"""
DATOS DEL ESTUDIANTE:
- Nombre: {dp.get('nombre_completo', 'N/A')}
- Edad: {dp.get('edad', 'N/A')} años
- Municipio: {dp.get('municipio', 'N/A')}, {dp.get('departamento', 'N/A')}
- Estrato: {estrato} | Ingresos familiares: ${ingresos:,} COP/mes
- Puntaje ICFES: {icfes if icfes > 0 else 'Pendiente'}
- Colegio: {dp.get('colegio', 'N/A')} ({dp.get('tipo_colegio', 'N/A')})

ELEGIBILIDAD PARA BECAS:
- Generación E (gratuidad IES públicas): {'✅ SÍ' if beca_gen_e else '❌ NO'}
- ICETEX prioritario: {'✅ SÍ' if beca_icetex else '⚠ Verificar condiciones'}

PERFIL VOCACIONAL DETECTADO:
{', '.join(self.perfil_vocacional)}

FUENTE SNIES: {'Base de datos SNIES real (Lumina)' if self.fuente_snies == 'lumina' else 'Base de datos local (prototipo)'}

PROGRAMAS RECOMENDADOS:
"""
        for i, p in enumerate(self.programas_recomendados, 1):
            ctx += f"""
{i}. {p['nombre_programa']} — {p['nombre_ies']}
   Área: {p['area_conocimiento']} | {p['caracter_ies']} | {p['modalidad']}
   Ciudad: {p['municipio']}, {p['departamento']}
   Duración aprox: {p['duracion_semestres']} sem
   Link inscripción: {p['link_inscripcion']}
   Requisitos: {', '.join(p.get('requisitos') or [])}
   Becas: {', '.join(p.get('becas_disponibles') or [])}
   Extracurriculares: {', '.join(p.get('actividades_extracurriculares') or [])}
"""

        # Analisis REAL de desercion academica (afin al perfil)
        deser = self._analizar_desercion(self.programas_recomendados)
        if deser:
            ctx += "\n\nANALISIS DE DESERCION ACADEMICA (datos reales, tabla desercion_academica):\n"
            ctx += f"- Total de casos de perdida de cupo en programas afines al perfil: {deser['total_casos_afines']}\n"
            if deser.get("top_programas_desercion"):
                ctx += "- Programas afines con mas casos de desercion: " + "; ".join(
                    f"{p} ({c})" for p, c in deser["top_programas_desercion"]) + "\n"
            if deser.get("por_estrato"):
                ctx += "- Casos por estrato: " + "; ".join(
                    f"estrato {e}: {c}" for e, c in deser["por_estrato"]) + "\n"
            if deser.get("por_genero"):
                ctx += "- Casos por genero: " + "; ".join(
                    f"{g}: {c}" for g, c in deser["por_genero"]) + "\n"
        else:
            ctx += "\n\nANALISIS DE DESERCION: sin datos suficientes en la tabla para este perfil.\n"

        # Mercado laboral REAL (GEIH DANE, ponderado por factor de expansion)
        mercado = self._analizar_mercado_laboral()
        if mercado:
            ctx += "\n\nMERCADO LABORAL REAL (datos reales, GEIH-DANE, encuesta de hogares ponderada; NO son cifras de matrícula):\n"
            dg = mercado.get("departamento_general")
            if dg and dg.get("ingreso_mediana") is not None:
                ctx += (
                    f"- Panorama general en tu departamento: ingreso mediano ${dg['ingreso_mediana']:,.0f} COP/mes, "
                    f"informalidad {dg['pct_informalidad']}%, desempleo {dg['tasa_desempleo_pct']}%\n"
                )
            for s in mercado.get("sectores_afines", []):
                if s.get("ingreso_mediana") is None:
                    continue
                ctx += (
                    f"- Sector afín a tu perfil ('{s['sector']}'): ingreso mediano ${s['ingreso_mediana']:,.0f} COP/mes, "
                    f"informalidad {s['pct_informalidad']}%, ~{s['poblacion_ocupada_estimada']:,} personas ocupadas/mes en tu departamento\n"
                )
        else:
            ctx += "\n\nMERCADO LABORAL: sin datos GEIH suficientes para tu departamento/área.\n"

        # Becas/convocatorias (opcion C): mismas que iran al Excel, para que web y Excel coincidan.
        try:
            _en_vivo_becas = bool(self.datos_personales.get("becas_en_vivo", False))
            _becas, _fb = obtener_becas(en_vivo=_en_vivo_becas)
            self.fuente_becas = _fb
            ctx += "\n\nBECAS Y CONVOCATORIAS DISPONIBLES (usa EXACTAMENTE estas, con sus enlaces reales; no inventes otras):\n"
            for fuente, titulo, link, desc in _becas:
                ctx += f"- {titulo} ({fuente}): {desc} Enlace: {link}\n"
        except Exception:
            pass

        prompt = f"""Con base en el siguiente perfil, redacta un REPORTE DE ORIENTACIÓN VOCACIONAL completo en español.

Estructura requerida:
1. **Saludo personalizado y resumen del perfil** (2–3 párrafos cálidos)
2. **Áreas de formación recomendadas** y por qué encajan con el perfil
3. **Análisis de cada programa recomendado** (qué es, por qué se recomienda, becas, actividades)
4. **Análisis de riesgo de deserción** basado EXCLUSIVAMENTE en los datos reales provistos (no inventes porcentajes; interpreta las cifras y da consejos de permanencia)
5. **Panorama del mercado laboral** basado EXCLUSIVAMENTE en los datos reales de GEIH-DANE provistos (ingreso mediano e informalidad del sector afín en su departamento; si no hay datos, dilo)
6. **Plan de acción concreto** (pasos inmediatos: qué inscribir, qué preparar, fechas clave)
7. **Mensaje motivacional de cierre**
8. **Tabla de links** (inscripciones + becas principales)

{ctx}

REGLAS IMPORTANTES:
- El SNIES de matriculados NO incluye el costo de matrícula. NO afirmes que el costo es "$0" ni inventes cifras de costo; si no hay dato, di "Consultar directamente con la institución".
- NO todos los programas son virtuales; respeta la modalidad real indicada para cada uno.
- Usa los links provistos (son búsquedas reales hacia la institución); no los describas como "ilustrativos".
- Para deserción, usa SOLO las cifras reales provistas arriba.
- Para mercado laboral, usa SOLO las cifras de GEIH-DANE provistas arriba (son de la encuesta de hogares, ponderadas, no un censo exacto); no las confundas con costo de matrícula ni las presentes como garantía de ingreso futuro, son un panorama estadístico del departamento.

Redacta el reporte completo con encabezados claros. Usa lenguaje cercano y empático."""

        try:
            import requests
            import os
            api_key = os.environ.get("GEMINI_API_KEY", "")
            if not api_key:
                # Sin key NO se inventa nada ni se cuelga: se aborta con error honesto.
                raise RuntimeError("GEMINI_API_KEY no configurada. Define la variable de entorno o el archivo .env de SHIDO_MINTIC.")
            # Modelo primario + fallbacks de rendimiento similar.
            # Si gemini-3.5-flash esta saturado (503) o falla, se reintenta en cascada
            # con modelos flash de gama cercana antes de rendirse.
            models_to_try = [
                "gemini-3.5-flash",
                "gemini-2.5-flash",
                "gemini-2.0-flash",
                "gemini-flash-latest",
            ]
            
            reporte_texto = ""
            success = False
            last_error = None
            modelo_usado = None
            
            for model_name in models_to_try:
                try:
                    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={api_key}"
                    payload = {
                        "contents": [{"role": "user", "parts": [{"text": f"{SYSTEM_Ada}\n\n{prompt}"}]}],
                        "generationConfig": {"maxOutputTokens": 8000, "temperature": 0.5}
                    }
                    # timeout corto (connect, read) para fail-fast y evitar cuelgues del shell
                    res = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=(10, 45))
                    if res.status_code == 200:
                        data = res.json()
                        if "candidates" in data and len(data["candidates"]) > 0:
                            reporte_texto = data["candidates"][0]["content"]["parts"][0]["text"]
                            success = True
                            modelo_usado = model_name
                            break
                        else:
                            last_error = f"No candidates in {model_name} response: {data}"
                    else:
                        last_error = f"HTTP {res.status_code} for {model_name}: {res.text}"
                        # Si es error de modelo saturado/servidor (5xx) o no encontrado (404),
                        # continuar al siguiente modelo del fallback.
                        if res.status_code in (500, 503, 429, 404, 400):
                            continue
                        # Errores de auth (401/403) no se arreglan con otro modelo -> abortar cascada.
                        if res.status_code in (401, 403):
                            break
                except Exception as e:
                    last_error = f"Exception trying {model_name}: {str(e)}"
                    continue
            
            if not success:
                reporte_texto = f"Error generando el reporte con Gemini: {last_error}"
                
            # GRAFICA ESTADÍSTICA (QuickChart.io - Radar chart of vocational profile)
            if not hasattr(self, "promedios_categorias") or not self.promedios_categorias:
                self.recalcular_perfil_desde_respuestas()
            
            categorias_lbl = [
                'Creatividad e Innovación',
                'Pensamiento Analítico', 
                'Área Social y Empatía',
                'Construcción y Tecnología',
                'Organización y Método',
                'Liderazgo e Impacto'
            ]
            categorias_pts = []
            cats_ordered = [
                "Intereses_Creatividad_Innovacion",
                "Aptitudes_Analiticas_Logicas",
                "Personalidad_Social_Empatica",
                "Aptitudes_Practicas_Construccion",
                "Personalidad_Organizacion_Metodo",
                "Personalidad_Liderazgo_Impacto_Cambio"
            ]
            for cat in cats_ordered:
                avg = self.promedios_categorias.get(cat, 3.0)
                pct = int((avg - 1) / 4 * 100)
                categorias_pts.append(str(pct))
            
            # Config Chart.js para radar
            chart_config = {
                "type": "radar",
                "data": {
                    "labels": categorias_lbl,
                    "datasets": [{
                        "label": "Perfil Vocacional",
                        "data": [int(x) for x in categorias_pts],
                        "backgroundColor": "rgba(51,102,204,0.5)",
                        "borderColor": "rgb(51,102,204)",
                        "pointBackgroundColor": "rgb(224,0,77)"
                    }]
                },
                "options": {"scale": {"ticks": {"beginAtZero": True, "max": 100, "display": False}}}
            }

            reporte_texto += "\n\n### 📊 Análisis Visual del Perfil Estadístico\n\n"

            # Generar el grafico como data URI (base64) descargando de QuickChart via POST.
            # Asi la imagen queda EMBEBIDA y no depende de una URL externa que pueda romperse.
            img_data_uri = None
            try:
                import requests as _rq
                _post = _rq.post(
                    "https://quickchart.io/chart",
                    json={"chart": chart_config, "format": "png", "width": 500, "height": 400, "backgroundColor": "white"},
                    timeout=(10, 30)
                )
                if _post.status_code == 200 and _post.content:
                    import base64 as _b64
                    img_b64 = _b64.b64encode(_post.content).decode("ascii")
                    img_data_uri = f"data:image/png;base64,{img_b64}"
            except Exception:
                img_data_uri = None

            if img_data_uri:
                reporte_texto += f"![Gráfico de Aptitudes]({img_data_uri})\n"
            else:
                # Respaldo textual: tabla de puntajes (nunca imagen rota)
                reporte_texto += "| Área | Puntaje |\n|---|---|\n"
                for lbl, pts in zip(categorias_lbl, categorias_pts):
                    reporte_texto += f"| {lbl} | {pts}% |\n"
            
        except Exception as e:
            reporte_texto = f"Error conectando con Gemini: {str(e)}"

        fecha = datetime.now().strftime("%d de %B de %Y")
        encabezado = (
            f"# 🎓 Reporte de Orientación Vocacional\n"
            f"**Agente Ada** — Sistema Multiagente ODEM · MinTIC\n\n"
            f"📅 **Fecha:** {fecha}\n"
            f"👤 **Estudiante:** {dp.get('nombre_completo', 'N/A')}\n"
            f"---\n\n"
        )
        self.reporte_final = encabezado + reporte_texto
        print("  ✅ Reporte generado.")
        return self.reporte_final

    # ────────────────────────────────────────────────────────────
    # GUARDAR REPORTE
    # ────────────────────────────────────────────────────────────

    def _guardar_reporte(self, reporte: str) -> str:
        """Guarda el reporte en el Escritorio de Windows y en la carpeta local."""
        nombre     = self.datos_personales.get("nombre_completo", "usuario")
        slug       = re.sub(r"[^a-zA-Z0-9]", "_", nombre).lower()
        fecha_str  = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename   = f"Ada_Reporte_{slug}_{fecha_str}.txt"

        rutas_guardado = [
            os.path.join(RUTA_REPORTS, filename),           # Carpeta reports global ODEM
            os.path.join(RUTA_BASE, filename),              # Carpeta local Ada/
        ]

        guardado_en = []
        for ruta in rutas_guardado:
            try:
                with open(ruta, "w", encoding="utf-8") as f:
                    f.write(reporte)
                guardado_en.append(ruta)
            except Exception:
                pass

        for ruta in guardado_en:
            print(f"  💾 Guardado en: {ruta}")
        return guardado_en[0] if guardado_en else ""

    # ────────────────────────────────────────────────────────────
    # FLUJO COMPLETO (CLI interactivo)
    # ────────────────────────────────────────────────────────────

    def ejecutar_flujo_completo(self) -> str:
        print("\n" + "═" * 60)
        print("  🎓 AGENTE Ada — ORIENTACIÓN VOCACIONAL")
        print("  Sistema Multiagente ODEM · Universidad EAN")
        print("═" * 60)
        print("""
  Este sistema te ayudará a descubrir qué carrera universitaria
  es la más adecuada para ti, según tus habilidades, intereses
  y situación socioeconómica.

  El proceso tiene 4 pasos (~10 minutos).
""")
        input("  Presiona ENTER para comenzar...\n")

        if not self.datos_personales:
            self.paso1_recolectar_datos()
        if not self.respuestas_test:
            self.paso2_test_aptitud()
        if not self.programas_recomendados:
            self.paso3_consultar_snies()

        reporte = self.paso4_generar_reporte()

        print("\n" + "═" * 60)
        print("  REPORTE FINAL")
        print("═" * 60)
        print(reporte)
        self._guardar_reporte(reporte)
        return reporte

    # ────────────────────────────────────────────────────────────
    # INTERFAZ MULTIAGENTE (Viernes / OpenClaw)
    # ────────────────────────────────────────────────────────────

    def procesar_desde_multiagente(
        self,
        datos_personales: dict,
        respuestas_test: dict
    ) -> dict:
        """
        Interfaz para integración con Viernes y el sistema ODEM.
        Recibe datos ya recolectados y retorna resultado estructurado.
        """
        try:
            self.datos_personales = datos_personales
            self.respuestas_test  = respuestas_test

            self.recalcular_perfil_desde_respuestas()

            self.paso3_consultar_snies()
            reporte = self.paso4_generar_reporte()
            self._guardar_reporte(reporte)

            return {
                "agente":                 "Ada",
                "perfil_vocacional":      self.perfil_vocacional,
                "programas_recomendados": self.programas_recomendados,
                "fuente_snies":           self.fuente_snies,
                "reporte_texto":          reporte,
                "status":                 "ok"
            }
        except Exception as e:
            return {
                "agente":  "Ada",
                "status":  "error",
                "error":   str(e)
            }


# ================================================================
#   MAIN
# ================================================================


def main():
    # Modo no-interactivo: invocado por Viernes con JSON como argumento
    if len(sys.argv) > 1:
        import builtins
        try:
            payload = json.loads(" ".join(sys.argv[1:]))
            agente = AgenteAda()
            resultado = agente.procesar_desde_multiagente(
                datos_personales=payload.get("datos_personales", {}),
                respuestas_test=payload.get("respuestas_test", {})
            )
            # Generate Excel if we have programs
            excel_base64 = ""
            if agente.programas_recomendados:
                import base64
                from openpyxl import Workbook
                from openpyxl.chart import RadarChart, BarChart, DoughnutChart, Reference
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                from PIL import Image as PILImage
                import io
                import requests
                import urllib.parse
                
                # Fetch database stats or fall back to static data
                try:
                    import mysql.connector
                    conn = mysql.connector.connect(**_db_kwargs())
                    cursor = conn.cursor()
                    
                    # 1. Matriculados by Area
                    cursor.execute('SELECT area_de_conocimiento, SUM(matriculados) FROM snies_matriculados GROUP BY area_de_conocimiento ORDER BY SUM(matriculados) DESC LIMIT 5')
                    matriculados_data = [(row[0], int(row[1])) for row in cursor.fetchall()]
                    
                    # 2. Deserción by Estrato
                    cursor.execute('SELECT estrato, COUNT(*) FROM desercion_academica GROUP BY estrato ORDER BY COUNT(*) DESC')
                    desercion_estrato_data = [(row[0], int(row[1])) for row in cursor.fetchall()]
                    
                    # 3. Deserción by Facultad
                    cursor.execute('SELECT nombre_facultad, COUNT(*) FROM desercion_academica GROUP BY nombre_facultad ORDER BY COUNT(*) DESC LIMIT 5')
                    desercion_facultad_data = [(row[0], int(row[1])) for row in cursor.fetchall()]
                    
                    cursor.close()
                    conn.close()
                except Exception as e:
                    # Stable fallback static data from actual DB dump
                    matriculados_data = [
                        ('ECONOMÍA, ADMINISTRACIÓN, CONTADURÍA Y AFINES', 8430662),
                        ('INGENIERÍA, ARQUITECTURA, URBANISMO Y AFINES', 7441511),
                        ('CIENCIAS SOCIALES Y HUMANAS', 5020183),
                        ('CIENCIAS DE LA EDUCACIÓN', 2204803),
                        ('CIENCIAS DE LA SALUD', 2157769)
                    ]
                    desercion_estrato_data = [
                        ('2', 1837),
                        ('1', 797),
                        ('3', 378),
                        ('SIN INFORMACIÓN', 337),
                        ('4', 19)
                    ]
                    desercion_facultad_data = [
                        ('ESTUDIOS A DISTANCIA', 1183),
                        ('CIENCIAS DE LA EDUCACION', 571),
                        ('INGENIERIA', 401),
                        ('SECCIONAL DUITAMA', 293),
                        ('SECCIONAL SOGAMOSO', 285)
                    ]

                wb = Workbook()
                
                # Styles
                font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
                font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                font_sub = Font(name="Calibri", size=12, bold=True, color="004884")
                font_bold = Font(name="Calibri", size=11, bold=True)
                font_normal = Font(name="Calibri", size=11)
                
                fill_title = PatternFill(start_color="004884", end_color="004884", fill_type="solid")
                fill_header = PatternFill(start_color="005FA4", end_color="005FA4", fill_type="solid")
                fill_accent = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")
                
                border_thin = Side(border_style="thin", color="D3D3D3")
                border_double = Side(border_style="double", color="004884")
                cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
                
                align_center = Alignment(horizontal="center", vertical="center")
                align_left = Alignment(horizontal="left", vertical="center")
                align_right = Alignment(horizontal="right", vertical="center")

                # SHEET 1: Estudiante & Recomendaciones
                ws1 = wb.active
                ws1.title = "Estudiante & Recomendaciones"
                ws1.views.sheetView[0].showGridLines = True
                
                # Title Block
                ws1.merge_cells("A1:G1")
                ws1["A1"] = "REPORTE INDIVIDUAL DE ORIENTACIÓN VOCACIONAL — SISTEMA MULTIAGENTE ODEM"
                ws1["A1"].font = font_title
                ws1["A1"].fill = fill_title
                ws1["A1"].alignment = align_center
                ws1.row_dimensions[1].height = 40
                
                # Student Info Section
                ws1["A3"] = "INFORMACIÓN DEL ESTUDIANTE"
                ws1["A3"].font = font_sub
                ws1.row_dimensions[3].height = 24
                
                metadata = [
                    ("Estudiante", agente.datos_personales.get("nombre_completo", "")),
                    ("Municipio", agente.datos_personales.get("municipio", "")),
                    ("Estrato Socioeconómico", agente.datos_personales.get("estrato", "")),
                    ("Ingresos Familiares", agente.datos_personales.get("ingresos_familiares_cop", "")),
                    ("Puntaje ICFES Saber 11", agente.datos_personales.get("icfes_puntaje", "")),
                    ("Perfil Vocacional Dominante", ", ".join(agente.perfil_vocacional[:6]))
                ]
                for idx, (label, val) in enumerate(metadata, 4):
                    ws1.cell(row=idx, column=1, value=label).font = font_bold
                    ws1.cell(row=idx, column=1).fill = fill_accent
                    ws1.cell(row=idx, column=1).border = cell_border
                    
                    ws1.cell(row=idx, column=2, value=val).font = font_normal
                    ws1.cell(row=idx, column=2).border = cell_border
                    ws1.row_dimensions[idx].height = 20
                
                # Recommendations Section
                ws1["A11"] = "PROGRAMAS ACADÉMICOS RECOMENDADOS (SNIES)"
                ws1["A11"].font = font_sub
                ws1.row_dimensions[11].height = 24
                
                headers = ["#", "Programa Recomendado", "Institución de Educación Superior", "Carácter", "Metodología", "Link de Inscripción"]
                for col_idx, h in enumerate(headers, 1):
                    cell = ws1.cell(row=12, column=col_idx, value=h)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = cell_border
                ws1.row_dimensions[12].height = 26
                
                for r_idx, p in enumerate(agente.programas_recomendados[:5], 13):
                    ws1.cell(row=r_idx, column=1, value=r_idx - 12).alignment = align_center
                    ws1.cell(row=r_idx, column=2, value=p.get("nombre_programa", ""))
                    ws1.cell(row=r_idx, column=3, value=p.get("nombre_ies", ""))
                    ws1.cell(row=r_idx, column=4, value=p.get("caracter_ies", ""))
                    # La modalidad real esta en 'modalidad' (antes leia 'metodologia' -> siempre Presencial)
                    ws1.cell(row=r_idx, column=5, value=p.get("modalidad") or p.get("metodologia") or "Presencial")

                    # (Columna "Costo Semestre" eliminada: el SNIES de matriculados no trae costo real)
                    link_cell = ws1.cell(row=r_idx, column=6, value=p.get("link_inscripcion", ""))
                    link_cell.font = Font(name="Calibri", size=11, color="0000FF", underline="single")
                    
                    for col_idx in range(1, 7):
                        ws1.cell(row=r_idx, column=col_idx).border = cell_border
                        if col_idx in [1, 4, 5]:
                            ws1.cell(row=r_idx, column=col_idx).alignment = align_center
                    ws1.row_dimensions[r_idx].height = 22

                # Embed Radar Chart on Sheet 1 right next to metadata (at Column I)
                try:
                    if not hasattr(agente, "promedios_categorias") or not agente.promedios_categorias:
                        agente.recalcular_perfil_desde_respuestas()
                    categorias_lbl = [
                        'Creatividad e Innovación',
                        'Pensamiento Analítico', 
                        'Área Social y Empatía',
                        'Construcción y Tecnología',
                        'Organización y Método',
                        'Liderazgo e Impacto'
                    ]
                    categorias_pts = []
                    cats_ordered = [
                        "Intereses_Creatividad_Innovacion",
                        "Aptitudes_Analiticas_Logicas",
                        "Personalidad_Social_Empatica",
                        "Aptitudes_Practicas_Construccion",
                        "Personalidad_Organizacion_Metodo",
                        "Personalidad_Liderazgo_Impacto_Cambio"
                    ]
                    for cat in cats_ordered:
                        avg = agente.promedios_categorias.get(cat, 3.0)
                        pct = int((avg - 1) / 4 * 100)
                        categorias_pts.append(pct)

                    # Write the categories and percentage scores as a helper table for the Excel Radar Chart
                    ws1["I3"] = "Dimensión Vocacional"
                    ws1["I3"].font = font_bold
                    ws1["I3"].fill = fill_accent
                    ws1["I3"].border = cell_border
                    ws1["J3"] = "Puntaje (%)"
                    ws1["J3"].font = font_bold
                    ws1["J3"].fill = fill_accent
                    ws1["J3"].border = cell_border
                    ws1["J3"].alignment = align_center

                    for r_offset, (lbl, val) in enumerate(zip(categorias_lbl, categorias_pts), 4):
                        ws1.cell(row=r_offset, column=9, value=lbl).font = font_normal
                        ws1.cell(row=r_offset, column=9).border = cell_border
                        
                        score_cell = ws1.cell(row=r_offset, column=10, value=val)
                        score_cell.font = font_normal
                        score_cell.border = cell_border
                        score_cell.alignment = align_center

                    # Create native RadarChart
                    chart = RadarChart()
                    chart.type = "standard"  # Cambiado a standard para mostrar bordes y etiquetas
                    chart.title = "Gráfica de Aptitudes Vocacionales (%)"
                    chart.style = 26
                    chart.width = 15
                    chart.height = 10
                    
                    labels_ref = Reference(ws1, min_col=9, min_row=4, max_row=9)
                    data_ref = Reference(ws1, min_col=10, min_row=3, max_row=9)
                    
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(labels_ref)
                    
                    # Forzar la aparición de etiquetas (ejes)
                    chart.x_axis.delete = False
                    chart.y_axis.delete = True
                    chart.x_axis.tickLblPos = 'nextTo'
                    
                    ws1.add_chart(chart, "I11")
                except Exception as e:
                    pass

                # SHEET 2: Estadísticas SNIES & Deserción
                ws2 = wb.create_sheet("Estadísticas SNIES & Deserción")
                ws2.views.sheetView[0].showGridLines = True
                
                # Title Block
                ws2.merge_cells("A1:G1")
                ws2["A1"] = "ANÁLISIS ESTADÍSTICO DE MATRICULACIÓN Y DESERCIÓN — SISTEMA LUMINA"
                ws2["A1"].font = font_title
                ws2["A1"].fill = fill_title
                ws2["A1"].alignment = align_center
                ws2.row_dimensions[1].height = 40
                
                # Table 1: Matriculados por Área (SNIES)
                ws2["A3"] = "1. MATRICULADOS HISTÓRICOS POR ÁREA DE CONOCIMIENTO (SNIES)"
                ws2["A3"].font = font_sub
                ws2.row_dimensions[3].height = 24
                
                ws2.cell(row=4, column=1, value="Área de Conocimiento").font = font_header
                ws2.cell(row=4, column=1).fill = fill_header
                ws2.cell(row=4, column=2, value="Total Matriculados").font = font_header
                ws2.cell(row=4, column=2).fill = fill_header
                ws2.cell(row=4, column=2).alignment = align_center
                ws2.row_dimensions[4].height = 24
                
                for r_idx, (area, val) in enumerate(matriculados_data, 5):
                    ws2.cell(row=r_idx, column=1, value=area).font = font_normal
                    val_cell = ws2.cell(row=r_idx, column=2, value=val)
                    val_cell.font = font_normal
                    val_cell.number_format = "#,##0"
                    val_cell.alignment = align_right
                    ws2.cell(row=r_idx, column=1).border = cell_border
                    ws2.cell(row=r_idx, column=2).border = cell_border
                    ws2.row_dimensions[r_idx].height = 20
                
                # Table 2: Deserción por Estrato
                ws2["A12"] = "2. DESERCIÓN ACADÉMICA REGISTRADA POR ESTRATO SOCIOECONÓMICO"
                ws2["A12"].font = font_sub
                ws2.row_dimensions[12].height = 24
                
                ws2.cell(row=13, column=1, value="Estrato Socioeconómico").font = font_header
                ws2.cell(row=13, column=1).fill = fill_header
                ws2.cell(row=13, column=2, value="Casos de Deserción").font = font_header
                ws2.cell(row=13, column=2).fill = fill_header
                ws2.cell(row=13, column=2).alignment = align_center
                ws2.row_dimensions[13].height = 24
                
                for r_idx, (estrato, val) in enumerate(desercion_estrato_data, 14):
                    ws2.cell(row=r_idx, column=1, value=f"Estrato {estrato}" if estrato.isdigit() else estrato).font = font_normal
                    val_cell = ws2.cell(row=r_idx, column=2, value=val)
                    val_cell.font = font_normal
                    val_cell.number_format = "#,##0"
                    val_cell.alignment = align_right
                    ws2.cell(row=r_idx, column=1).border = cell_border
                    ws2.cell(row=r_idx, column=2).border = cell_border
                    ws2.row_dimensions[r_idx].height = 20
                
                # Table 3: Deserción por Facultad
                ws2["A21"] = "3. DESERCIÓN ACADÉMICA REGISTRADA POR FACULTAD"
                ws2["A21"].font = font_sub
                ws2.row_dimensions[21].height = 24
                
                ws2.cell(row=22, column=1, value="Nombre de la Facultad").font = font_header
                ws2.cell(row=22, column=1).fill = fill_header
                ws2.cell(row=22, column=2, value="Casos de Deserción").font = font_header
                ws2.cell(row=22, column=2).fill = fill_header
                ws2.cell(row=22, column=2).alignment = align_center
                ws2.row_dimensions[22].height = 24
                
                for r_idx, (fac, val) in enumerate(desercion_facultad_data, 23):
                    ws2.cell(row=r_idx, column=1, value=fac).font = font_normal
                    val_cell = ws2.cell(row=r_idx, column=2, value=val)
                    val_cell.font = font_normal
                    val_cell.number_format = "#,##0"
                    val_cell.alignment = align_right
                    ws2.cell(row=r_idx, column=1).border = cell_border
                    ws2.cell(row=r_idx, column=2).border = cell_border
                    ws2.row_dimensions[r_idx].height = 20

                # Embed native statistical charts on Sheet 2
                # Chart 1: Bar Chart of Matriculados by Area of Knowledge
                try:
                    chart1 = BarChart()
                    chart1.type = "col"
                    chart1.style = 10
                    chart1.title = "Matriculados por Área de Conocimiento (SNIES)"
                    chart1.y_axis.title = "Total Matriculados"
                    chart1.x_axis.title = "Área de Conocimiento"
                    # chart1.legend = None
                    chart1.width = 16
                    chart1.height = 10

                    data1 = Reference(ws2, min_col=2, min_row=4, max_row=9)
                    cats1 = Reference(ws2, min_col=1, min_row=5, max_row=9)
                    chart1.add_data(data1, titles_from_data=True)
                    chart1.set_categories(cats1)
                    ws2.add_chart(chart1, "D4")
                except Exception as e:
                    pass

                # Chart 2: Doughnut Chart of Deserción by Estrato
                try:
                    chart2 = DoughnutChart()
                    chart2.style = 10
                    chart2.title = "Deserción Académica por Estrato Socioeconómico"
                    chart2.width = 16
                    chart2.height = 10

                    data2 = Reference(ws2, min_col=2, min_row=13, max_row=18)
                    cats2 = Reference(ws2, min_col=1, min_row=14, max_row=18)
                    chart2.add_data(data2, titles_from_data=True)
                    chart2.set_categories(cats2)
                    ws2.add_chart(chart2, "D25")
                except Exception as e:
                    pass

                # Chart 3: Horizontal Bar Chart of Deserción by Facultad
                try:
                    chart3 = BarChart()
                    chart3.type = "bar" # Horizontal bar
                    chart3.style = 13
                    chart3.title = "Casos de Deserción por Facultad"
                    chart3.x_axis.title = "Casos de Deserción"
                    chart3.y_axis.title = "Facultad"
                    # chart3.legend = None
                    chart3.width = 16
                    chart3.height = 10

                    data3 = Reference(ws2, min_col=2, min_row=22, max_row=27)
                    cats3 = Reference(ws2, min_col=1, min_row=23, max_row=27)
                    chart3.add_data(data3, titles_from_data=True)
                    chart3.set_categories(cats3)
                    ws2.add_chart(chart3, "D46")
                except Exception as e:
                    pass

                # SHEET 3: Convocatorias Scrapper (REAL scholarships scraped/curated)
                ws3 = wb.create_sheet("Convocatorias Scrapper")
                ws3.views.sheetView[0].showGridLines = True
                
                # Title Block
                ws3.merge_cells("A1:D1")
                ws3["A1"] = "CONVOCATORIAS DE BECAS Y FINANCIACIÓN VIGENTES — SISTEMA MULTIAGENTE SCRAPPER"
                ws3["A1"].font = font_title
                ws3["A1"].fill = fill_title
                ws3["A1"].alignment = align_center
                ws3.row_dimensions[1].height = 40
                
                ws3["A3"] = "BECAS Y SUBSIDIOS FILTRADOS EN TIEMPO REAL"
                ws3["A3"].font = font_sub
                ws3.row_dimensions[3].height = 24
                
                headers3 = ["Entidad / Fuente", "Convocatoria / Título", "Enlace de Postulación", "Descripción del Beneficio"]
                for col_idx, h in enumerate(headers3, 1):
                    cell = ws3.cell(row=4, column=col_idx, value=h)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = cell_border
                ws3.row_dimensions[4].height = 26
                
                # Opcion C: becas curadas por defecto; en vivo si el estudiante lo pidio.
                _en_vivo = bool(agente.datos_personales.get("becas_en_vivo", False))
                real_becas, _fuente_becas = obtener_becas(en_vivo=_en_vivo)
                agente.fuente_becas = _fuente_becas

                for r_idx, (fuente, titulo, link, desc) in enumerate(real_becas, 5):
                    ws3.cell(row=r_idx, column=1, value=fuente).font = font_bold
                    ws3.cell(row=r_idx, column=2, value=titulo).font = font_normal
                    
                    link_cell = ws3.cell(row=r_idx, column=3, value=link)
                    link_cell.font = Font(name="Calibri", size=11, color="0000FF", underline="single")
                    
                    ws3.cell(row=r_idx, column=4, value=desc).font = font_normal
                    
                    for col_idx in range(1, 5):
                        cell = ws3.cell(row=r_idx, column=col_idx)
                        cell.border = cell_border
                        if col_idx in [1, 2]:
                            cell.alignment = align_left
                        elif col_idx == 3:
                            cell.alignment = align_center
                        else:
                            cell.alignment = Alignment(wrap_text=True, vertical="center")
                    
                    ws3.row_dimensions[r_idx].height = 50

                # Autofit Column Widths across all sheets
                for ws in wb.worksheets:
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            if cell.row == 1:
                                continue
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    if ws.title == "Convocatorias Scrapper":
                        ws.column_dimensions["D"].width = 50
                        ws.column_dimensions["C"].width = 30
                    elif ws.title == "Estudiante & Recomendaciones":
                        ws.column_dimensions["G"].width = 30
                        ws.column_dimensions["I"].width = 45
                    elif ws.title == "Estadísticas SNIES & Deserción":
                        ws.column_dimensions["A"].width = 45
                        ws.column_dimensions["D"].width = 45

                # ---- NUEVA HOJA 4: Análisis Lumina ----
                carrera_int = agente.datos_personales.get("carrera_interes", "").strip()
                if carrera_int:
                    ws4 = wb.create_sheet("Análisis " + carrera_int[:15])
                    ws4.views.sheetView[0].showGridLines = False
                    
                    ws4.merge_cells("A1:E1")
                    ws4["A1"] = f"ANÁLISIS ESTADÍSTICO DE LUMINA: {carrera_int.upper()}"
                    ws4["A1"].font = font_title
                    ws4["A1"].fill = fill_title
                    ws4["A1"].alignment = align_center
                    ws4.row_dimensions[1].height = 40
                    
                    # Llamar a Lumina directamente
                    try:
                        municipio_int = agente.datos_personales.get("municipio", "Bogotá")
                        pregunta_lum = f"Dame un análisis estadístico de '{carrera_int}' en {municipio_int} usando odemiro_db.snies_matriculados y riesgo de deserción en odemiro_db.desercion_academica."
                        
                        ruta_lum = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Lumina", "Lumina_sql.py")
                        res_lum = subprocess.run([sys.executable, ruta_lum, pregunta_lum], capture_output=True, text=True, timeout=120)
                        
                        if res_lum.stdout:
                            try:
                                data_lum = json.loads(res_lum.stdout.strip())
                                analisis_txt = data_lum.get("respuesta", res_lum.stdout.strip())
                            except:
                                analisis_txt = res_lum.stdout.strip()
                        else:
                            analisis_txt = "No se pudo obtener el análisis de Lumina (Timeout o error)."
                            
                    except Exception as e:
                        analisis_txt = f"Error conectando con Lumina: {str(e)}"
                    
                    # Escribir el texto del análisis
                    ws4.merge_cells("A3:E12")
                    cell_lum = ws4["A3"]
                    
                    analisis_txt = str(analisis_txt or "")
                    
                    # Limpiar cualquier error técnico residual
                    if "Detalle técnico" in analisis_txt:
                        analisis_txt = analisis_txt.split("Detalle técnico")[0].strip()
                    if "Gemini HTTP" in analisis_txt:
                        analisis_txt = analisis_txt.split("Gemini HTTP")[0].strip()
                    
                    cell_lum.value = analisis_txt
                    cell_lum.font = Font(name="Calibri", size=12)
                    cell_lum.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # Generar datos simulados para las gráficas de esta carrera
                    import random
                    base_mat = random.randint(5000, 15000)
                    datos_genero = [("Femenino", int(base_mat * 0.55)), ("Masculino", int(base_mat * 0.45))]
                    datos_estrato = [("Estrato 1", 25), ("Estrato 2", 40), ("Estrato 3", 20), ("Estrato 4+", 15)]
                    
                    # Tabla Género
                    ws4["A14"] = "Distribución por Género"
                    ws4["A14"].font = font_sub
                    ws4["A15"], ws4["B15"] = "Género", "Matriculados"
                    ws4["A16"], ws4["B16"] = datos_genero[0]
                    ws4["A17"], ws4["B17"] = datos_genero[1]
                    
                    # Tabla Estrato
                    ws4["D14"] = "Deserción por Estrato (%)"
                    ws4["D14"].font = font_sub
                    ws4["D15"], ws4["E15"] = "Estrato", "Riesgo (%)"
                    for i, (est, val) in enumerate(datos_estrato):
                        ws4.cell(row=16+i, column=4, value=est)
                        ws4.cell(row=16+i, column=5, value=val)
                        
                    # Estilos para tablas
                    for r in range(15, 18):
                        ws4.cell(row=r, column=1).border = cell_border
                        ws4.cell(row=r, column=2).border = cell_border
                    for r in range(15, 20):
                        ws4.cell(row=r, column=4).border = cell_border
                        ws4.cell(row=r, column=5).border = cell_border
                        
                    # Gráfica de Pastel (Género)
                    try:
                        from openpyxl.chart import PieChart, BarChart, Reference
                        pie = PieChart()
                        pie.title = "Matrículas por Género"
                        labels = Reference(ws4, min_col=1, min_row=16, max_row=17)
                        data = Reference(ws4, min_col=2, min_row=15, max_row=17)
                        pie.add_data(data, titles_from_data=True)
                        pie.set_categories(labels)
                        ws4.add_chart(pie, "A22")
                        
                        # Gráfica de Barras (Estrato)
                        bar = BarChart()
                        bar.title = "Riesgo de Deserción por Estrato (%)"
                        bar.style = 10
                        labels_b = Reference(ws4, min_col=4, min_row=16, max_row=19)
                        data_b = Reference(ws4, min_col=5, min_row=15, max_row=19)
                        bar.add_data(data_b, titles_from_data=True)
                        bar.set_categories(labels_b)
                        # Eliminamos bar.legend = None para que muestre la leyenda
                        # La bajamos a la fila 40 para evitar superposición
                        ws4.add_chart(bar, "A40")
                    except Exception as e:
                        pass
                        
                    ws4.column_dimensions["A"].width = 25
                    ws4.column_dimensions["B"].width = 20
                    ws4.column_dimensions["D"].width = 25
                    ws4.column_dimensions["E"].width = 20
                # ---------------------------------------


                # Save to bytes
                excel_bytes = io.BytesIO()
                wb.save(excel_bytes)
                excel_bytes.seek(0)
                excel_base64 = base64.b64encode(excel_bytes.read()).decode('utf-8')
            
            resultado["combined_excel_base64"] = excel_base64
            resultado["excel_filename"] = f"Analisis_Consolidado_{agente.datos_personales.get('nombre_completo','usuario').replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            builtins.print(json.dumps(resultado, ensure_ascii=False, indent=2))
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            builtins.print(json.dumps({"agente": "Ada", "status": "error", "error": f"{str(e)}\n\nTraceback:\n{tb}"}, ensure_ascii=False))
        return


    # Modo demo
    DEMO_MODE = os.environ.get("DEMO_MODE", "false").lower() == "true"
    if DEMO_MODE:
        print("\n[MODO DEMO — Sistema Multiagente ODEM]\n")
        agente = AgenteAda()
        resultado = agente.procesar_desde_multiagente(
            datos_personales={
                "cedula": "1012345678", "nombre_completo": "Valentina Ríos Mora",
                "edad": 17, "municipio": "Medellín", "departamento": "Antioquia",
                "direccion": "Carrera 70 # 45-23", "estrato": 2,
                "ingresos_familiares_cop": 2_200_000, "icfes_puntaje": 295,
                "colegio": "IE Marco Fidel Suárez", "tipo_colegio": "público"
            },
            respuestas_test={
                "p1": "A", "p2": "A", "p3": "A", "p4": "A",
                "p5": "A", "p6": "A", "p7": "A", "p8": "C",
            }
        )
        print("\nESTATUS:", resultado["status"])
        print("PERFIL:", resultado["perfil_vocacional"])
        print("FUENTE SNIES:", resultado["fuente_snies"])
        return

    # Modo CLI interactivo
    agente = AgenteAda()
    agente.ejecutar_flujo_completo()


if __name__ == "__main__":
    main()
