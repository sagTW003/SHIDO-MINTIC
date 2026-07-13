# ================================================================
#   SCRAPER UNIVERSAL DE CURSOS EN LINEA  v7.0
#   Platzi - Udemy - Coursera - DataCamp - Domestika - EanX
# ================================================================
#
#  INSTALACION (una sola vez en CMD):
#      pip install selenium webdriver-manager openpyxl
#
#  ABRIR CHROME EN MODO DEBUG (CMD como Administrador):
#  "C:/Program Files/Google/Chrome/Application/chrome.exe"
#  --remote-debugging-port=9222
#  --user-data-dir="C:/Users/TU_USER/chrome_scraper"
#
#  Luego navega a la pagina de listado de cursos y ejecuta:
#      python course_scraper.py
# ================================================================

import time, re, sys, os, random
from datetime import datetime

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        TimeoutException, ElementClickInterceptedException, WebDriverException
    )
    from webdriver_manager.chrome import ChromeDriverManager
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Ejecuta: pip install selenium webdriver-manager openpyxl")
    sys.exit(1)


# ================================================================
# PERFILES: selectores verificados por la estructura real de cada
# plataforma. "links" = como encontrar los <a> en el LISTADO.
# "det_*"  = como extraer cada campo en la pagina de DETALLE.
# ================================================================
PERFILES = {
    "platzi.com": {
        "nombre": "Platzi",
        "precio_default": "Suscripcion",
        # El listado de Platzi muestra tarjetas como <a href="/cursos/slug/">
        "links": ["a[href*='/cursos/'][href$='/']", "a[href*='/cursos/']"],
        "det_titulo":       ["h1", "h1[class*='Title']", "h2[class*='title']"],
        "det_desc":         ["[class*='SummaryContent'] p",
                             "[class*='CourseContent'] p",
                             "[class*='description'] p",
                             "section p", "main p"],
        "det_rating":       ["[class*='Rating']", "[class*='rating']",
                             "[aria-label*='stars']"],
        "det_cert":         ["[class*='Certificate']", "[class*='certificate']"],
        "det_precio":       [],   # suscripcion
    },

    "udemy.com": {
        "nombre": "Udemy",
        "precio_default": None,
        # Los cards de Udemy envuelven el titulo en un <a href="/course/slug/">
        "links": ["a[href*='/course/'][class*='link']",
                  "a[href*='/course/']"],
        "det_titulo":       ["h1[data-purpose='lead-title']", "h1"],
        "det_desc":         ["[data-purpose='course-description'] p",
                             "[class*='description--content'] p",
                             "[class*='about-section'] p"],
        "det_rating":       ["[data-purpose='rating-number']",
                             "span[class*='rating-number']"],
        "det_cert":         ["[class*='certificate']",
                             "[class*='Certificate']"],
        "det_precio":       ["[data-purpose='course-price-text'] [class*='price-part']",
                             "span[class*='price-text--price-part']",
                             "[class*='base-price']"],
    },

    "coursera.org": {
        "nombre": "Coursera",
        "precio_default": None,
        "links": ["a[href*='/learn/']", "a[href*='/specializations/']",
                  "a[href*='/professional-certificates/']"],
        "det_titulo":       ["h1[class*='title']", "h1[class*='cds']", "h1"],
        "det_desc":         ["[class*='description'] p",
                             "[class*='about-section'] p",
                             "[data-testid='description'] p"],
        "det_rating":       ["[class*='ratings-text']",
                             "[aria-label*='stars']",
                             "[class*='RatingsCount']"],
        "det_cert":         ["[class*='CertificateSection']",
                             "[class*='certificate']",
                             "[aria-label*='certificate']"],
        "det_precio":       ["[class*='price']", "[aria-label*='price']"],
    },

    "datacamp.com": {
        "nombre": "DataCamp",
        "precio_default": "Suscripcion",
        "links": ["a[href*='/courses/']", "a[href*='/tracks/']",
                  "[class*='course-block'] a"],
        "det_titulo":       ["h1", "h2[class*='title']"],
        "det_desc":         ["[class*='course-description'] p",
                             "[class*='description'] p",
                             "section p"],
        "det_rating":       ["[class*='rating']"],
        "det_cert":         ["[class*='certificate']", "[class*='badge']"],
        "det_precio":       [],
    },

    "domestika.org": {
        "nombre": "Domestika",
        "precio_default": None,
        "links": ["a[href*='/courses/']", "[class*='course-in-list'] a",
                  "article a[href*='/courses/']"],
        "det_titulo":       ["h1[class*='title']", "h1"],
        "det_desc":         ["[class*='course-description'] p",
                             "[class*='description'] p",
                             "section p"],
        "det_rating":       ["[class*='average__number']",
                             "[class*='score']"],
        "det_cert":         ["[class*='certificate']"],
        "det_precio":       ["[class*='course-price'] [class*='amount']",
                             "[class*='price'] [class*='amount']",
                             "[class*='price']"],
    },

    "eanx.co": {
        "nombre": "EanX",
        "precio_default": None,
        "links": ["a[href*='/course/']", "a[href*='/cursos/']",
                  ".course-card a", "article a"],
        "det_titulo":       ["h1", "h2[class*='title']",
                             "[class*='page-title']"],
        "det_desc":         ["[class*='summary'] p",
                             "[class*='description'] p",
                             "#summary p", "section p"],
        "det_rating":       ["[class*='rating']", "[class*='stars']"],
        "det_cert":         ["[class*='certificate']"],
        "det_precio":       ["[class*='price']", "[class*='costo']"],
    },
}

GENERICO = {
    "nombre": "Desconocida",
    "precio_default": None,
    "links": ["a[href*='/course']", "a[href*='/curso']",
              "[class*='card'] a", "article a"],
    "det_titulo":   ["h1", "h2"],
    "det_desc":     ["[class*='description'] p", "main p", "p"],
    "det_rating":   ["[class*='rating']", "[class*='stars']"],
    "det_cert":     ["[class*='certificate']", "[class*='cert']"],
    "det_precio":   ["[class*='price']"],
}


# ================================================================
# UTILIDADES
# ================================================================

def detectar_perfil(url):
    for dominio, p in PERFILES.items():
        if dominio in url.lower():
            return p
    return GENERICO


def espera_humana(min_s=1.5, max_s=4.0):
    """Pausa aleatoria para imitar comportamiento humano."""
    time.sleep(random.uniform(min_s, max_s))


def hay_captcha(driver):
    """Detecta si hay un CAPTCHA activo en la pagina."""
    indicadores = [
        "iframe[src*='recaptcha']",
        "iframe[src*='captcha']",
        "iframe[src*='hcaptcha']",
        "[class*='captcha']",
        "[id*='captcha']",
    ]
    for sel in indicadores:
        if driver.find_elements(By.CSS_SELECTOR, sel):
            return True
    # Tambien buscar texto tipico de captcha en el body
    try:
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        if "captcha" in body and len(body) < 2000:
            return True
    except Exception:
        pass
    return False


def esperar_captcha_manual(driver, url_actual):
    """
    Si se detecta CAPTCHA, pausa el script y espera que
    el usuario lo resuelva manualmente en Chrome.
    """
    print("\n  ⚠  CAPTCHA detectado en:", url_actual[:60])
    print("  → Resuelvelo manualmente en Chrome y luego presiona Enter aqui.")
    input("  → [Enter para continuar] ")
    # Esperar que la pagina recargue tras resolver
    WebDriverWait(driver, 20).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )
    espera_humana(4, 6)


def extraer_css(driver, selectores):
    """Prueba selectores CSS en orden y retorna el primer texto encontrado."""
    for sel in selectores:
        try:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            textos = [e.text.strip() for e in els if e.text.strip()]
            if textos:
                return " ".join(textos[:3])[:600]  # max 3 parrafos, 600 chars
        except Exception:
            continue
    return "N/A"


def extraer_xpath(driver, xpath):
    """Extrae texto de un XPath dado por el usuario."""
    try:
        el = driver.find_element(By.XPATH, xpath)
        return el.text.strip()[:600] or "N/A"
    except Exception:
        return "N/A"


# ================================================================
# PASO 1 — RECOLECTAR LINKS DEL LISTADO
# ================================================================

def _clic_seguro(driver, elemento):
    """Clica un elemento, usando JS si el clic normal falla."""
    try:
        elemento.click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", elemento)


def _btn_siguiente(driver, xpath_boton):
    """
    Busca el boton de avanzar pagina. Prioridad:
    1. XPath dado por el usuario.
    2. Patrones conocidos de paginacion numerada (boton ">", "Next", etc.).
    Retorna el elemento si existe y es clicable, None si no.
    """
    # XPath del usuario (boton cargar mas o siguiente pagina)
    if xpath_boton:
        try:
            return WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, xpath_boton))
            )
        except TimeoutException:
            pass

    # Patrones de paginacion numerada — selectores CSS y XPaths comunes
    selectores_siguiente = [
        # Flechas y textos comunes
        "a[aria-label='Next']", "a[aria-label='Siguiente']",
        "button[aria-label='Next']", "button[aria-label='Siguiente']",
        "[class*='next']:not([disabled])", "[class*='Next']:not([disabled])",
        "[rel='next']",
        # XPaths para texto visible
        "//a[normalize-space()='>']", "//button[normalize-space()='>']",
        "//a[normalize-space()='›']", "//button[normalize-space()='›']",
        "//a[contains(text(),'Next')]", "//button[contains(text(),'Next')]",
        "//a[contains(text(),'Siguiente')]",
        "//li[contains(@class,'next')]/a",
        "//li[not(contains(@class,'disabled'))]/a[contains(@class,'next')]",
    ]

    for sel in selectores_siguiente:
        try:
            by = By.XPATH if sel.startswith("//") else By.CSS_SELECTOR
            btn = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable((by, sel))
            )
            # Verificar que no este deshabilitado
            disabled = btn.get_attribute("disabled") or btn.get_attribute("aria-disabled")
            clase    = btn.get_attribute("class") or ""
            if disabled or "disabled" in clase.lower():
                continue
            return btn
        except TimeoutException:
            continue

    return None


def recolectar_links(driver, perfil, cantidad, xpath_boton, url_base):
    """
    Recolecta links de cursos soportando los 3 mecanismos de carga:
      1. Paginacion numerada  (< 1 2 3 ... N >)
      2. Boton 'Cargar mas'
      3. Scroll infinito
    En cada pagina acumula los links nuevos hasta alcanzar `cantidad`.
    """
    print(f"\nRecolectando links (objetivo: {cantidad})...")
    todos_links = {}   # dict href -> True para deduplicar manteniendo orden
    sin_progreso = 0

    while True:
        if hay_captcha(driver):
            esperar_captcha_manual(driver, driver.current_url)

        # Acumular links de la pagina actual
        nuevos = _links_actuales(driver, perfil, url_base)
        antes  = len(todos_links)
        for lnk in nuevos:
            todos_links[lnk] = True
        print(f"  Links acumulados: {len(todos_links)}", end="\r")

        if len(todos_links) >= cantidad:
            break

        # ── Intentar avanzar (paginacion o cargar mas) ──────────
        btn = _btn_siguiente(driver, xpath_boton)
        if btn:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            espera_humana(4.5, 5)
            _clic_seguro(driver, btn)
            espera_humana(4, 5)
            # Esperar que la pagina actualice su contenido
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            sin_progreso = 0
            continue

        # ── Sin boton: intentar scroll infinito ─────────────────
        alto_antes = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        espera_humana(2, 3)
        alto_despues = driver.execute_script("return document.body.scrollHeight")

        if alto_despues == alto_antes:
            sin_progreso += 1
            if sin_progreso >= 3:
                print("\n  No hay mas contenido para cargar.")
                break
        else:
            sin_progreso = 0

    resultado = list(todos_links.keys())[:cantidad]
    print(f"\n  Total links encontrados: {len(resultado)}")
    return resultado


def _links_actuales(driver, perfil, url_base):
    """Extrae y filtra los href de los cursos visibles en la pagina."""
    dominio = re.sub(r"https?://", "", url_base).split("/")[0]
    vistos, links = set(), []

    for sel in perfil["links"]:
        try:
            for el in driver.find_elements(By.CSS_SELECTOR, sel):
                href = (el.get_attribute("href") or "").split("?")[0].rstrip("/")
                if href and href not in vistos and dominio in href:
                    vistos.add(href)
                    links.append(href)
        except Exception:
            continue

    # Fallback: todos los <a> del dominio que parezcan cursos
    if not links:
        keywords = ["/course", "/curso", "/learn", "/class"]
        excluir  = ["/blog", "/login", "/signup", "/about",
                    "/contact", "/pricing", "/terms", "/privacy"]
        try:
            for el in driver.find_elements(By.TAG_NAME, "a"):
                href = (el.get_attribute("href") or "").split("?")[0].rstrip("/")
                if (href and href not in vistos and dominio in href
                        and any(k in href for k in keywords)
                        and not any(x in href for x in excluir)):
                    vistos.add(href)
                    links.append(href)
        except Exception:
            pass

    return links


# ================================================================
# PASO 2 — EXTRAER DATOS DE CADA PAGINA DE DETALLE
# ================================================================

def extraer_detalle(driver, url, perfil, xpaths_user):
    """Navega a la URL del curso y extrae todos los campos."""
    try:
        driver.get(url)
        WebDriverWait(driver, 12).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1, h2"))
            )
        except TimeoutException:
            pass

        if hay_captcha(driver):
            esperar_captcha_manual(driver, url)

        espera_humana(4, 5.5)

        # Extraer cada campo: XPath del usuario tiene prioridad
        titulo = (extraer_xpath(driver, xpaths_user["titulo"])
                  if xpaths_user.get("titulo")
                  else extraer_css(driver, perfil["det_titulo"]))

        desc = (extraer_xpath(driver, xpaths_user["desc"])
                if xpaths_user.get("desc")
                else extraer_css(driver, perfil["det_desc"]))

        rating = (extraer_xpath(driver, xpaths_user["rating"])
                  if xpaths_user.get("rating")
                  else extraer_css(driver, perfil["det_rating"]))
        # Limpiar rating: solo numero
        if rating != "N/A":
            m = re.search(r"\d+[.,]\d+|\d+", rating)
            rating = m.group(0) if m else rating

        # Certificado: selector + keyword en el body
        cert_raw = (extraer_xpath(driver, xpaths_user["cert"])
                    if xpaths_user.get("cert")
                    else extraer_css(driver, perfil["det_cert"]))
        if cert_raw != "N/A":
            certificado = "Si"
        else:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            certificado = "Si" if any(
                k in body for k in ["certificado", "certificate", "diploma"]
            ) else "N/A"

        # Precio
        if perfil.get("precio_default"):
            precio = perfil["precio_default"]
        elif xpaths_user.get("precio"):
            precio = extraer_xpath(driver, xpaths_user["precio"])
        else:
            precio = extraer_css(driver, perfil["det_precio"])

        return {
            "Titulo":       titulo,
            "Descripcion":  desc,
            "Calificacion": rating,
            "Certificado":  certificado,
            "Precio":       precio,
            "URL":          url,
        }

    except Exception as e:
        return {"Titulo": "ERROR", "Descripcion": str(e)[:80],
                "Calificacion": "N/A", "Certificado": "N/A",
                "Precio": "N/A", "URL": url}


# ================================================================
# EXPORTAR EXCEL
# ================================================================

def exportar_excel(cursos, plataforma, url):
    if not cursos:
        return ""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = plataforma[:28]

    AZUL, BLANCO, CELESTE = "1A1A2E", "FFFFFF", "E8F0FE"
    borde = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = (f"Plataforma: {plataforma}  |  {url[:55]}  |  "
               f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  {len(cursos)} cursos")
    c.font = Font(bold=True, size=10, color=AZUL)
    c.fill = PatternFill("solid", fgColor=CELESTE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    headers = ["N°", "Titulo", "Descripcion", "Calificacion",
               "Certificado", "Precio", "URL"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color=BLANCO, size=11)
        c.fill      = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde
    ws.row_dimensions[2].height = 25

    campos = ["Titulo", "Descripcion", "Calificacion",
              "Certificado", "Precio", "URL"]
    for row, curso in enumerate(cursos, 3):
        fill = PatternFill("solid", fgColor="F0F4FF" if row % 2 == 0 else BLANCO)
        ws.cell(row=row, column=1, value=row-2).fill = fill
        ws.cell(row=row, column=1).border    = borde
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        for col, campo in enumerate(campos, 2):
            c = ws.cell(row=row, column=col, value=curso.get(campo, "N/A"))
            c.fill      = fill
            c.border    = borde
            c.alignment = Alignment(
                horizontal="center" if col in [4, 5, 6] else "left",
                vertical="top", wrap_text=True
            )
            if col == 5 and str(curso.get(campo)).lower() == "si":
                c.font = Font(bold=True, color="217346")
        ws.row_dimensions[row].height = 55

    for col, w in zip(range(1, 8), [5, 34, 48, 13, 13, 14, 40]):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes    = "A3"
    ws.auto_filter.ref = f"A2:G{len(cursos)+2}"

    nombre = (f"cursos_{re.sub(r'[^a-zA-Z0-9]','_', plataforma)}"
              f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(nombre)
    return nombre


# ================================================================
# CONEXION A CHROME
# ================================================================

def conectar_chrome():
    options = Options()
    options.add_experimental_option("debuggerAddress", "localhost:9222")
    print("Conectando a Chrome en puerto 9222...")
    try:
        service = Service(ChromeDriverManager().install())
        driver  = webdriver.Chrome(service=service, options=options)
        print(f"  OK - URL activa: {driver.current_url}\n")
        return driver
    except WebDriverException as e:
        err = str(e).lower()
        if "cannot connect" in err or "refused" in err:
            print("\nERROR: Chrome no esta corriendo en el puerto 9222.")
            print("Abre Chrome con el comando del encabezado de este archivo.")
        else:
            print(f"\nERROR: {e}")
        sys.exit(1)


# ================================================================
# INPUT DEL USUARIO
# ================================================================

def pedir(msg, requerido=True):
    while True:
        v = input(f"  > {msg}: ").strip()
        if v:
            return v
        if not requerido:
            return ""
        print("  Campo obligatorio.")


def recopilar():
    print("\n" + "=" * 60)
    print("  SCRAPER UNIVERSAL DE CURSOS EN LINEA  v7.0")
    print("=" * 60)
    print("""
FUNCIONAMIENTO:
  1. Recorre el listado haciendo scroll o usando boton 'Cargar mas'.
  2. Recolecta los links de cada curso.
  3. Entra a cada pagina de detalle y extrae los datos.
  4. Si aparece un CAPTCHA, el script pausa y te pide que lo
     resuelvas manualmente en Chrome. Luego continua solo.

Los XPaths de la pagina de DETALLE son opcionales para:
  Platzi, Udemy, Coursera, DataCamp, Domestika, EanX.
  Presiona Enter para usar deteccion automatica.

Como obtener un XPath:
  Clic derecho en el elemento > Inspeccionar >
  clic derecho en el HTML > Copiar > XPath completo
""")
    print("-" * 60)

    url      = pedir("URL del LISTADO de cursos")
    cantidad = int(pedir("Cantidad de cursos a extraer"))

    print("\n-- XPaths en la pagina de DETALLE (Enter = automatico) --")
    xpaths = {
        "titulo": pedir("XPath del TITULO          (Enter=auto)", requerido=False),
        "desc":   pedir("XPath de la DESCRIPCION   (Enter=auto)", requerido=False),
        "rating": pedir("XPath de la CALIFICACION  (Enter=auto)", requerido=False),
        "cert":   pedir("XPath del CERTIFICADO     (Enter=auto)", requerido=False),
        "precio": pedir("XPath del PRECIO          (Enter=auto)", requerido=False),
    }
    boton = pedir("XPath del boton 'Cargar mas' en el LISTADO (Enter=scroll)",
                  requerido=False)

    return url, cantidad, xpaths, boton


# ================================================================
# MAIN
# ================================================================

def main():
    url, cantidad, xpaths, boton = recopilar()

    perfil     = detectar_perfil(url)
    plataforma = perfil["nombre"]
    if plataforma == "Desconocida":
        plataforma = url.split("/")[2].replace("www.", "").split(".")[0].capitalize()
    print(f"\nPlataforma: {plataforma}")

    driver = conectar_chrome()

    # Navegar al listado si hace falta
    if url.rstrip("/") not in driver.current_url:
        print(f"Navegando a: {url}")
        driver.get(url)
        WebDriverWait(driver, 15).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        espera_humana(4, 5)

    # Paso 1: links
    links = recolectar_links(driver, perfil, cantidad, boton, url)
    if not links:
        print("\nNo se encontraron links. Verifica que estes en el listado de cursos.")
        return

    # Paso 2: detalle de cada curso
    print(f"\nExtrayendo datos de {len(links)} cursos...\n")
    cursos = []
    for i, link in enumerate(links, 1):
        print(f"  [{i:03d}/{len(links)}] {link[-55:]}")
        datos = extraer_detalle(driver, link, perfil, xpaths)
        datos["N"] = i
        cursos.append(datos)
        print(f"         Titulo: {datos['Titulo'][:55]}")
        espera_humana(4, 5)   # pausa humana entre cursos

    # Volver al listado
    try:
        driver.get(url)
    except Exception:
        pass

    # Paso 3: Excel
    nombre = exportar_excel(cursos, plataforma, url)
    if nombre:
        print("\n" + "=" * 60)
        print(f"  Cursos extraidos : {len(cursos)}")
        print(f"  Plataforma       : {plataforma}")
        print(f"  Archivo          : {nombre}")
        print(f"  Ubicacion        : {os.path.abspath(nombre)}")
        print("=" * 60 + "\n")


if __name__ == "__main__":
    main()