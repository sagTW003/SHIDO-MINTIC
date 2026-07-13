"""
============================================================
  SCRAPPER — Agente de Web Scraping y Minería de Datos
  Sistema Multiagente ODEM · Universidad EAN
  Versión: 3.0.0 (migrado a Playwright)

  Complementa a Lumina con datos actualizados de la web:
  - Fechas de inscripción en tiempo real
  - Convocatorias de becas y subsidios
  - Datos de datos.gov.co no disponibles en SNIES estático
  - Scraping de plataformas educativas (Platzi, Udemy, Coursera, etc.)

  COMPATIBILIDAD:
  - Linux (Raspberry Pi OS, Ubuntu, Debian)
  - Windows (WSL o nativo)
  - macOS

  DEPENDENCIAS:
    pip install openai playwright openpyxl
    python -m playwright install chromium

  En Raspberry Pi:
    sudo apt install chromium-browser
    python -m playwright install-deps chromium

  USO:
    python3 Scrapper.py                          # Modo interactivo
    python3 Scrapper.py "buscar becas ICETEX"    # Modo consulta IA
    python3 Scrapper.py '{"url":"...", ...}'     # Modo multiagente (JSON)
============================================================
"""

import openai
import time, re, sys, os, random, json, subprocess
from datetime import datetime

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Ejecuta: pip install playwright openpyxl && python -m playwright install chromium")
    sys.exit(1)

# ================================================================
#   CONFIGURACIÓN
# ================================================================

def _cargar_env():
    """Carga variables de un .env en la raiz de SHIDO_MINTIC (sin dependencias)."""
    aqui = os.path.dirname(os.path.abspath(__file__))
    raiz = os.path.dirname(os.path.dirname(os.path.dirname(aqui)))  # .../SHIDO_MINTIC
    ruta_env = os.path.join(raiz, ".env")
    if not os.path.exists(ruta_env):
        return
    try:
        with open(ruta_env, "r", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if not linea or linea.startswith("#") or "=" not in linea:
                    continue
                clave, _, valor = linea.partition("=")
                clave = clave.strip()
                valor = valor.strip().strip('"').strip("'")
                if clave and clave not in os.environ:
                    os.environ[clave] = valor
    except Exception:
        pass


_cargar_env()

# Motor IA: Gemini (leido del .env). NVIDIA queda como opcional/fallback.
_getenv = os.environ.get
GEMINI_API_KEY = _getenv("GEMINI_API_KEY", "")
GEMINI_MODEL   = "gemini-3.5-flash"
API_KEY = _getenv("NVIDIA_API_KEY", "")  # opcional (fallback)
MODEL   = "nvidia/nemotron-3-ultra-550b-a55b"

# Ruta del ejecutable Chromium en Raspberry Pi (si playwright no lo detecta)
CHROMIUM_RASPBERRY = "/usr/bin/chromium-browser"

# ================================================================
#   PERFILES DE PLATAFORMAS EDUCATIVAS
# ================================================================

PERFILES = {
    "platzi.com": {
        "nombre": "Platzi",
        "precio_default": "Suscripcion",
        "links": ["a[href*='/cursos/'][href$='/']", "a[href*='/cursos/']"],
        "det_titulo":   ["h1", "h1[class*='Title']", "h2[class*='title']"],
        "det_desc":     ["[class*='SummaryContent'] p", "[class*='CourseContent'] p", "section p"],
        "det_rating":   ["[class*='Rating']", "[class*='rating']"],
        "det_cert":     ["[class*='Certificate']", "[class*='certificate']"],
        "det_precio":   [],
    },
    "udemy.com": {
        "nombre": "Udemy",
        "precio_default": None,
        "links": ["a[href*='/course/'][class*='link']", "a[href*='/course/']"],
        "det_titulo":   ["h1[data-purpose='lead-title']", "h1"],
        "det_desc":     ["[data-purpose='course-description'] p", "[class*='description--content'] p"],
        "det_rating":   ["[data-purpose='rating-number']", "span[class*='rating-number']"],
        "det_cert":     ["[class*='certificate']"],
        "det_precio":   ["[data-purpose='course-price-text'] [class*='price-part']"],
    },
    "coursera.org": {
        "nombre": "Coursera",
        "precio_default": None,
        "links": ["a[href*='/learn/']", "a[href*='/specializations/']"],
        "det_titulo":   ["h1[class*='title']", "h1"],
        "det_desc":     ["[class*='description'] p", "[data-testid='description'] p"],
        "det_rating":   ["[class*='ratings-text']", "[aria-label*='stars']"],
        "det_cert":     ["[class*='CertificateSection']"],
        "det_precio":   ["[class*='price']"],
    },
    "datos.gov.co": {
        "nombre": "Datos Gov Co",
        "precio_default": "Gratuito",
        "links": ["a[href*='/dataset/']", "a[href*='/resource/']", ".dataset-title a"],
        "det_titulo":   ["h1", "h1[class*='title']", ".dataset-title"],
        "det_desc":     [".dataset-description p", ".notes p", "section p"],
        "det_rating":   [],
        "det_cert":     [],
        "det_precio":   [],
    },
    "icetex.gov.co": {
        "nombre": "ICETEX",
        "precio_default": "Beca/Crédito",
        "links": ["a[href*='/becas']", "a[href*='/creditos']", "a[href*='/convocatoria']", ".card a"],
        "det_titulo":   ["h1", "h2[class*='title']", ".page-title"],
        "det_desc":     ["[class*='description'] p", ".convocatoria p", "main p"],
        "det_rating":   [],
        "det_cert":     [],
        "det_precio":   ["[class*='valor']", "[class*='monto']"],
    },
    "eanx.co": {
        "nombre": "EanX",
        "precio_default": None,
        "links": ["a[href*='/course/']", "a[href*='/cursos/']", ".course-card a"],
        "det_titulo":   ["h1", "h2[class*='title']"],
        "det_desc":     ["[class*='summary'] p", "[class*='description'] p"],
        "det_rating":   ["[class*='rating']"],
        "det_cert":     ["[class*='certificate']"],
        "det_precio":   ["[class*='price']"],
    },
}

GENERICO = {
    "nombre": "Web",
    "precio_default": None,
    "links": ["a[href*='/course']", "a[href*='/curso']", "a[href*='/beca']",
              "a[href*='/convocatoria']", "[class*='card'] a", "article a"],
    "det_titulo":   ["h1", "h2"],
    "det_desc":     ["[class*='description'] p", "main p", "p"],
    "det_rating":   ["[class*='rating']"],
    "det_cert":     ["[class*='certificate']"],
    "det_precio":   ["[class*='price']", "[class*='precio']"],
}

# ================================================================
#   UTILIDADES
# ================================================================

def detectar_perfil(url: str) -> dict:
    for dominio, p in PERFILES.items():
        if dominio in url.lower():
            return p
    return GENERICO

def espera_humana(min_s=1.5, max_s=3.5):
    time.sleep(random.uniform(min_s, max_s))

def extraer_css(page, selectores: list) -> str:
    for sel in selectores:
        try:
            elements = page.query_selector_all(sel)
            textos = []
            for el in elements[:3]:
                t = el.inner_text().strip()
                if t:
                    textos.append(t)
            if textos:
                return " ".join(textos)[:600]
        except Exception:
            continue
    return "N/A"

# ================================================================
#   PLAYWRIGHT — BROWSER
# ================================================================

def _chromium_executable() -> str | None:
    """Detecta Chromium disponible en el sistema (para Raspberry Pi)."""
    candidatos = [
        CHROMIUM_RASPBERRY,
        "/usr/bin/chromium",
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
    ]
    for c in candidatos:
        if os.path.exists(c):
            return c
    return None

def lanzar_browser(playwright, headless: bool = False):
    """
    Lanza Chromium via Playwright.
    - headless=False: muestra el navegador (útil para pruebas)
    - headless=True: sin GUI (modo servidor / Raspberry Pi)
    Detecta Chromium del sistema si Playwright no tiene el suyo.
    """
    kwargs = {
        "headless": headless,
        "args": [
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
            "--disable-dev-shm-usage",
        ],
    }

    # En Raspberry Pi, usar el Chromium del sistema si está disponible
    sys_chrome = _chromium_executable()
    if sys_chrome:
        kwargs["executable_path"] = sys_chrome

    try:
        browser = playwright.chromium.launch(**kwargs)
    except Exception as e:
        if "headless" in str(e).lower() or "executable" in str(e).lower():
            # Fallback headless con chromium del sistema
            kwargs["headless"] = True
            browser = playwright.chromium.launch(**kwargs)
        else:
            raise

    ctx = browser.new_context(
        viewport={"width": 1280, "height": 800},
        user_agent=(
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        ),
        locale="es-CO",
    )
    page = ctx.new_page()
    return browser, ctx, page


def recolectar_links(page, perfil: dict, cantidad: int, url_base: str) -> list:
    """Recorre el listado haciendo scroll y recoge links únicos."""
    links = set()
    scroll_intentos = 0

    while len(links) < cantidad and scroll_intentos < 20:
        for sel in perfil["links"]:
            try:
                elements = page.query_selector_all(sel)
                for el in elements:
                    href = el.get_attribute("href") or ""
                    if href.startswith("http") and href != url_base:
                        links.add(href.split("?")[0].rstrip("/"))
                    elif href.startswith("/"):
                        from urllib.parse import urlparse
                        base = urlparse(url_base)
                        full = f"{base.scheme}://{base.netloc}{href.split('?')[0].rstrip('/')}"
                        links.add(full)
            except Exception:
                continue

        if len(links) >= cantidad:
            break

        page.evaluate("window.scrollBy(0, 1200)")
        espera_humana(2, 3)
        scroll_intentos += 1

    return list(links)[:cantidad]


def extraer_detalle(page, url: str, perfil: dict) -> dict:
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_load_state("networkidle", timeout=8000)
        espera_humana(1.5, 2.5)
    except Exception:
        return {
            "URL": url, "Titulo": "Error carga", "Descripcion": "N/A",
            "Calificacion": "N/A", "Certificado": "N/A", "Precio": "N/A",
            "Plataforma": perfil["nombre"],
            "FechaExtraccion": datetime.now().strftime("%Y-%m-%d"),
        }

    titulo = extraer_css(page, perfil["det_titulo"])
    desc   = extraer_css(page, perfil["det_desc"])
    rating = extraer_css(page, perfil["det_rating"])
    cert   = extraer_css(page, perfil["det_cert"])
    precio = (extraer_css(page, perfil["det_precio"])
              if perfil["det_precio"]
              else (perfil.get("precio_default") or "N/A"))

    return {
        "URL":             url,
        "Titulo":          titulo,
        "Descripcion":     desc,
        "Calificacion":    rating,
        "Certificado":     cert,
        "Precio":          precio,
        "Plataforma":      perfil["nombre"],
        "FechaExtraccion": datetime.now().strftime("%Y-%m-%d"),
    }

# ================================================================
#   EXPORTAR EXCEL
# ================================================================

def exportar_excel(cursos: list, plataforma: str, url_base: str) -> str:
    ts         = datetime.now().strftime("%Y%m%d_%H%M")
    slug       = re.sub(r"[^a-zA-Z0-9]", "_", plataforma)[:20]
    carpeta    = os.path.dirname(os.path.abspath(__file__))
    # Guardar en la carpeta reports global en vez de quemar el Desktop
    reports_dir = os.path.abspath(os.path.join(carpeta, "..", "..", "..", "reports"))
    nombre     = f"Scrapper_{slug}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    encabezados = ["#", "Plataforma", "Titulo", "Descripcion", "Calificacion",
                   "Certificado", "Precio", "FechaExtraccion", "URL"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for i, curso in enumerate(cursos, 2):
        ws.cell(row=i, column=1,  value=curso.get("N", i - 1))
        ws.cell(row=i, column=2,  value=curso.get("Plataforma", ""))
        ws.cell(row=i, column=3,  value=curso.get("Titulo", ""))
        ws.cell(row=i, column=4,  value=curso.get("Descripcion", ""))
        ws.cell(row=i, column=5,  value=curso.get("Calificacion", ""))
        ws.cell(row=i, column=6,  value=curso.get("Certificado", ""))
        ws.cell(row=i, column=7,  value=curso.get("Precio", ""))
        ws.cell(row=i, column=8,  value=curso.get("FechaExtraccion", ""))
        ws.cell(row=i, column=9,  value=curso.get("URL", ""))
        if i % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="EFF6FF")

    anchos = [5, 14, 45, 60, 14, 14, 14, 16, 50]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # Guardar en reports y carpeta local
    rutas = [os.path.join(reports_dir, nombre), os.path.join(carpeta, nombre)]
    guardado = None
    for ruta in rutas:
        try:
            wb.save(ruta)
            guardado = ruta
            print(f"  💾 Excel guardado: {ruta}")
        except Exception:
            pass
    return guardado or nombre

# ================================================================
#   MÓDULO IA — Análisis con Claude
# ================================================================

historial = []

SYSTEM_SCRAPPER = (
    "Eres Scrapper, agente de web scraping y minería de datos del sistema multiagente ODEM "
    "(Universidad EAN). Tu función es recolectar información actualizada de la web para "
    "complementar los datos estáticos del SNIES que maneja Lumina. "
    "Especializaciones: fechas de inscripción en tiempo real, convocatorias de becas y subsidios, "
    "datasets de datos.gov.co, cursos y programas en plataformas educativas. "
    "Cuando Viernes o MINTIC te deleguen una tarea de scraping, ejecutas la recolección, "
    "analizas los datos y devuelves un JSON estructurado con los resultados. "
    "Eres preciso, eficiente y orientado al impacto social."
)

def _chat_gemini(mensaje: str) -> str:
    """Llama a Gemini via REST. Incluye system prompt + historial. Lanza excepcion si falla."""
    import requests
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY no configurada (.env de SHIDO_MINTIC).")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
    contents = []
    es_primero = True
    for msg in historial + [{"role": "user", "content": mensaje}]:
        rol = "user" if msg["role"] == "user" else "model"
        texto = msg["content"]
        if es_primero and rol == "user":
            texto = f"INSTRUCCIONES DEL SISTEMA:\n{SYSTEM_SCRAPPER}\n\nMENSAJE:\n{texto}"
            es_primero = False
        contents.append({"role": rol, "parts": [{"text": texto}]})
    payload = {"contents": contents, "generationConfig": {"temperature": 0.3, "maxOutputTokens": 4096}}
    res = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=(10, 60))
    if res.status_code != 200:
        raise RuntimeError(f"Gemini HTTP {res.status_code}: {res.text[:200]}")
    data = res.json()
    cand = data["candidates"][0]
    partes = cand.get("content", {}).get("parts", [])
    return "".join(p.get("text", "") for p in partes).strip()


def chat_ia(client, mensaje: str) -> str:
    """Motor IA del Scrapper: Gemini por defecto; NVIDIA como fallback si hay key."""
    respuesta = None
    try:
        respuesta = _chat_gemini(mensaje)
    except Exception as e_gem:
        # Fallback a NVIDIA solo si hay cliente y key
        if client is not None and API_KEY:
            try:
                messages_to_send = [{"role": "system", "content": SYSTEM_SCRAPPER}] + historial
                messages_to_send.append({"role": "user", "content": mensaje})
                response = client.chat.completions.create(
                    model=MODEL, max_tokens=2048, messages=messages_to_send
                )
                respuesta = response.choices[0].message.content
            except Exception:
                raise e_gem
        else:
            raise e_gem
    historial.append({"role": "user", "content": mensaje})
    historial.append({"role": "assistant", "content": respuesta})
    return respuesta

def analizar_resultados(client, cursos: list, objetivo: str) -> str:
    resumen = json.dumps(cursos[:20], ensure_ascii=False)
    prompt = (
        f"Objetivo del scraping: {objetivo}\n\n"
        f"Datos recolectados ({len(cursos)} items):\n{resumen}\n\n"
        f"Analiza estos resultados y genera un resumen útil con insights clave, "
        f"tendencias y recomendaciones para orientación educativa o análisis de mercado."
    )
    return chat_ia(client, prompt)

# ================================================================
#   MAIN
# ================================================================

def main():
    # Cliente NVIDIA opcional (solo fallback). El motor principal es Gemini.
    client = None
    if API_KEY:
        try:
            client = openai.OpenAI(api_key=API_KEY, base_url="https://integrate.api.nvidia.com/v1")
        except Exception:
            client = None

    # ── Modo no-interactivo: JSON desde multiagente (Viernes/MINTIC) ──
    if len(sys.argv) > 1:
        raw = " ".join(sys.argv[1:])
        try:
            payload = json.loads(raw)
            # Modo scraping estructurado desde multiagente
            url      = payload.get("url", "")
            cantidad = int(payload.get("cantidad", 10))
            objetivo = payload.get("objetivo", "scraping ODEM")
            headless = payload.get("headless", True)

            if url:
                perfil = detectar_perfil(url)
                cursos = []
                with sync_playwright() as pw:
                    browser, ctx, page = lanzar_browser(pw, headless=headless)
                    page.goto(url, wait_until="domcontentloaded", timeout=20000)
                    try:
                        page.wait_for_load_state("networkidle", timeout=15000)
                    except PWTimeout:
                        pass
                    espera_humana(3, 4)
                    links = recolectar_links(page, perfil, cantidad, url)
                    for i, link in enumerate(links, 1):
                        datos = extraer_detalle(page, link, perfil)
                        datos["N"] = i
                        cursos.append(datos)
                        espera_humana(2, 4)
                    browser.close()

                analisis = analizar_resultados(client, cursos, objetivo)
                archivo  = exportar_excel(cursos, perfil["nombre"], url)
                print(json.dumps({
                    "agente": "Scrapper", "tipo": "SCRAPING",
                    "plataforma": perfil["nombre"], "total": len(cursos),
                    "archivo_excel": archivo, "analisis": analisis,
                    "cursos": cursos,
                }, ensure_ascii=False))
            else:
                # Modo consulta IA pura
                respuesta = chat_ia(client, payload.get("mensaje", raw))
                print(json.dumps({
                    "agente": "Scrapper", "tipo": "CHAT", "respuesta": respuesta
                }, ensure_ascii=False))
        except json.JSONDecodeError:
            # Arg simple = consulta IA en texto plano
            respuesta = chat_ia(client, raw)
            print(json.dumps({
                "agente": "Scrapper", "tipo": "CHAT", "respuesta": respuesta
            }, ensure_ascii=False))
        return

    # ── Modo interactivo ──────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  SCRAPPER — Agente de Web Scraping · ODEM · Universidad EAN")
    print("  Motor: Playwright (Chromium) — compatible Raspberry Pi")
    print("=" * 60)
    print("\nModos disponibles:")
    print("  1. Scraping de cursos (Platzi, Udemy, Coursera, EanX, etc.)")
    print("  2. Consulta IA (becas, fechas, convocatorias, análisis)")
    print("  3. Salir")

    while True:
        modo = input("\n  Modo (1/2/3): ").strip()

        if modo == "3":
            break

        elif modo == "2":
            print("\n  Modo consulta IA. Escribe 'salir' para volver.")
            while True:
                pregunta = input("  Tú: ").strip()
                if pregunta.lower() == "salir":
                    break
                respuesta = chat_ia(client, pregunta)
                print(f"  Scrapper: {respuesta}\n")

        elif modo == "1":
            url      = input("  URL del listado de cursos: ").strip()
            cantidad = int(input("  Cantidad de cursos a extraer (ej. 5): ").strip() or "5")
            objetivo = input("  Objetivo del scraping: ").strip()
            headless_input = input("  ¿Modo silencioso sin ventana? (s/n) [s]: ").strip().lower()
            headless = headless_input != "n"

            perfil = detectar_perfil(url)
            print(f"\n  Plataforma detectada: {perfil['nombre']}")
            print(f"  Motor: Playwright Chromium | headless={headless}\n")

            cursos = []
            try:
                with sync_playwright() as pw:
                    browser, ctx, page = lanzar_browser(pw, headless=headless)

                    print(f"  Navegando a: {url}")
                    page.goto(url, wait_until="domcontentloaded", timeout=20000)
                    try:
                        page.wait_for_load_state("networkidle", timeout=10000)
                    except PWTimeout:
                        pass
                    espera_humana(3, 5)

                    print("  Recolectando links...")
                    links = recolectar_links(page, perfil, cantidad, url)
                    print(f"  {len(links)} links encontrados.")

                    for i, link in enumerate(links, 1):
                        print(f"  [{i:03d}/{len(links)}] {link[-60:]}")
                        datos = extraer_detalle(page, link, perfil)
                        datos["N"] = i
                        cursos.append(datos)
                        espera_humana(2, 4)

                    browser.close()

            except Exception as e:
                print(f"\n  ❌ Error durante el scraping: {e}")
                if cursos:
                    print(f"  Se recolectaron {len(cursos)} items antes del error.")
                else:
                    continue

            if cursos:
                print("\n  Analizando resultados con IA...")
                analisis = analizar_resultados(client, cursos, objetivo)
                print(f"\n  Scrapper: {analisis}")
                exportar_excel(cursos, perfil["nombre"], url)
            else:
                print("  No se recolectaron datos.")

        else:
            print("  Opción no válida.")


if __name__ == "__main__":
    main()
